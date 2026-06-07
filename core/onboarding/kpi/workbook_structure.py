"""Structural workbook reader for KPI files.

Reads an .xlsx into the (columns, rows, merged_spans) triple the format detector
needs. Merged-cell ranges are the structural signal that markitdown/CSV flattening
destroys — a label cell that spans several rows is how a parent KPI visually
brackets its sub-KPIs — so they must be read from the workbook itself, before any
flattening, and handed to ``detect_kpi_format(merged_spans=...)``.

Workspace-agnostic: it reads only structure (headers, cell values, merge ranges),
never domain content. openpyxl is already a project dependency (polars uses it for
.xlsx); a missing engine degrades to "no merge info", not an error.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class WorkbookGrid:
    """A flattened sheet plus the merge structure that flattening would lose.

    - ``columns``      : header strings (first non-empty row)
    - ``rows``         : list of dict rows keyed by header (data rows after header)
    - ``merged_spans`` : (col_index, row_start, row_end) per merged range, both
                         indices 0-based into the DATA rows (header excluded). A
                         span with ``row_end > row_start`` is a vertical merge.
    """

    columns: list[str]
    rows: list[dict[str, Any]]
    merged_spans: list[tuple[int, int, int]]


def _first_sheet(wb: Any) -> Any:
    """Return the first sheet in document order.

    openpyxl's ``wb.active`` returns whichever sheet the file's internal
    metadata marks as "active", which varies across xlsx editors and save
    operations.  For a multi-sheet workbook this means two consecutive reads
    of the same file can hit different sheets, producing different KPI counts
    (the 3-vs-42 non-determinism).  Always using the first sheet in document
    order (``wb.worksheets[0]``) is deterministic regardless of editor or save
    history. This mirrors the XML fallback which hardcodes ``sheet1.xml``.
    """
    sheets = wb.worksheets
    if not sheets:
        return wb.active  # degenerate: trust openpyxl
    return sheets[0]


def read_merged_spans(path: str | Path, *, header_rows: int = 1) -> list[tuple[int, int, int]]:
    """Return merged-cell ranges as (col_index, row_start, row_end), 0-based into
    data rows (header excluded). Empty list if the file/engine can't be read."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    try:
        # Not read-only: merged-cell ranges are unavailable on read-only worksheets.
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return []
    try:
        ws = _first_sheet(wb)
        spans: list[tuple[int, int, int]] = []
        for rng in ws.merged_cells.ranges:
            col_index = rng.min_col - 1  # openpyxl is 1-based
            row_start = rng.min_row - 1 - header_rows
            row_end = rng.max_row - 1 - header_rows
            if row_end < 0:  # merge entirely within the header band
                continue
            spans.append((col_index, max(row_start, 0), row_end))
        return spans
    finally:
        wb.close()


def read_workbook_grid(path: str | Path, *, header_rows: int = 1) -> WorkbookGrid:
    """Read an .xlsx into columns + dict rows + merged spans for the detector.

    Always reads the first sheet in document order (index 0) rather than
    ``wb.active``, which is non-deterministic across xlsx editors.  See
    ``_first_sheet`` for the rationale.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = _first_sheet(wb)
        matrix = [
            ["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()
    if not matrix:
        return WorkbookGrid(columns=[], rows=[], merged_spans=[])
    columns = matrix[header_rows - 1] if header_rows >= 1 else matrix[0]
    columns = [c or f"col_{i}" for i, c in enumerate(columns)]
    rows: list[dict[str, Any]] = []
    for raw in matrix[header_rows:]:
        rows.append({columns[i]: (raw[i] if i < len(raw) else "") for i in range(len(columns))})
    return WorkbookGrid(
        columns=columns,
        rows=rows,
        merged_spans=read_merged_spans(path, header_rows=header_rows),
    )


__all__ = ["WorkbookGrid", "read_merged_spans", "read_workbook_grid"]
