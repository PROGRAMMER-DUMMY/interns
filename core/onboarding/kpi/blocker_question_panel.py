"""Build a stakeholder-friendly blocker question panel from KPI feature mapping."""
from __future__ import annotations
from core.observability.cost_ledger import anchored

import argparse
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from core.paths import PROJECT_ROOT
from core.profiling.dataset_identity import dataset_display_stem
from core.storage.workspace_layout import WorkspaceLayout
from core.wiki import WikiLayout, read_feature_note
from core.contracts.versioning import register_contract
from core.onboarding.kpi.panel_preview_cache import (
    compute_preview_cache_key,
    load_cached_preview,
    save_cached_preview,
)
from core.onboarding.kpi.panel_preview_executor import execute_preview
from core.governance.injection_guard import neutralize_json, neutralize_rows, neutralize_text
from core.onboarding.kpi.pii_redaction import is_pii_column, redact_rows, redact_sample_values


PANEL_VERSION = 1

# Intent-contract facets are normally advisory (surfaced in the panel SET, enforced
# via gate-provenance, not made `current`). The exception is a facet that ALSO hard-
# blocks the execution harness: it must be answerable as `current` so the operator
# can resolve it via apply-kpi-panel-answer instead of looping on an empty panel.
_HARD_BLOCKING_INTENT_FACETS = frozenset({"grain_bucketing"})

register_contract("blocker_question_panel/current.json", current_version=PANEL_VERSION)
INTERACTION_CONTRACT = {
    "display_mode": "project_blocker_panel",
    "primary_artifact": "current.md",
    "answer_source": "current.json",
    "generic_answer_picker_allowed": False,
    "preserve_panel_options_only": True,
    "instruction": (
        "Show current.md verbatim, or render options directly from current.json. "
        "Do not summarize the blocker panel and do not use a generic answer picker "
        "that adds options outside this artifact."
    ),
}
READY_STATES = {
    "proven_direct",
    "proven_alias",
    "proven_join",
    "proven_formula",
    "proven_taxonomy",
    "user_confirmed",
}


@dataclass(frozen=True)
class BlockerQuestionPanelResult:
    output_dir: str
    current_json: str
    current_markdown: str
    current_full_markdown: str
    index_json: str
    question_count: int
    current_feature: str

    def summary(self) -> dict[str, Any]:
        return asdict(self)


class BlockerQuestionPanelBuilder:
    def __init__(
        self,
        repo_root: str | Path,
        workspace: str | Path,
        *,
        mapping_path: str | Path | None = None,
        output_dir: str | Path | None = None,
        deferred_kpi_ids: set[str] | None = None,
    ) -> None:
        self.repo_root = Path(repo_root).resolve()
        self.workspace = (self.repo_root / workspace).resolve()
        self.layout = WorkspaceLayout(project_root=self.workspace)
        # Partial-completion: KPIs with no measurable definition (empty metric AND
        # grain) are DEFERRED, not blocked. Their unresolved feature tokens must
        # NOT create feature-blocker questions -- otherwise a defined KPI is held
        # hostage by an undefined sibling's features. The id set comes from the
        # caller (the flow owns the definition gate); if not supplied, the builder
        # self-derives it from the KPI registry so it is correct standalone.
        if deferred_kpi_ids is None:
            deferred_kpi_ids = _deferred_kpi_ids_from_registry(self.workspace)
        self.deferred_kpi_ids = set(deferred_kpi_ids or set())
        self.mapping_path = (
            (self.repo_root / mapping_path).resolve()
            if mapping_path
            else self.layout.contracts_dir / "kpi_feature_mapping.json"
        )
        self.output_dir = (
            (self.repo_root / output_dir).resolve()
            if output_dir
            else self.layout.reports_dir / "blocker_question_panel"
        )

    def run(self) -> BlockerQuestionPanelResult:
        if not self.mapping_path.exists():
            raise FileNotFoundError(f"feature mapping not found: {self.mapping_path}")
        mapping = json.loads(self.mapping_path.read_text(encoding="utf-8"))
        self.output_dir.mkdir(parents=True, exist_ok=True)

        feature_questions = _build_questions(
            mapping, self.workspace, self.repo_root, self.deferred_kpi_ids
        )
        # Route low-confidence KPI intent-contract facets into the panel SET
        # (index.json) so they are visible/answerable. Most are advisory -- surfaced
        # + enforced via gate-provenance (--require-human-gates), not hard-blocking
        # the KPI flow -- so they stay out of `current`. Additive; never breaks
        # panel emission.
        intent_questions: list[dict[str, Any]] = []
        try:
            from core.onboarding.kpi.intent_contract import intent_facet_panel_questions

            intent_questions = intent_facet_panel_questions(
                self.repo_root, _rel(self.workspace, self.repo_root)
            )
        except Exception:  # pragma: no cover - defensive; intent routing is additive
            intent_questions = []
        # Near-tied base (fact) table selections from the source-to-target plan.
        # These HARD-block generation (the planner marks the KPI blocked), so
        # they must be answerable here via apply-kpi-panel-answer.
        base_source_questions: list[dict[str, Any]] = []
        try:
            from core.onboarding.relationships.base_source_selector import (
                base_source_panel_questions,
            )

            base_source_questions = base_source_panel_questions(
                self.repo_root, _rel(self.workspace, self.repo_root)
            )
        except Exception:  # pragma: no cover - defensive; routing is additive
            base_source_questions = []
        # `current` (the answerable / flow-blocking panel) is an unresolved
        # feature-mapping cluster first, preserving flow-stop semantics. Next a
        # near-tied base-source choice (it blocks the KPI's plan outright). With
        # none left, fall back to the first HARD-blocking intent facet
        # (grain_bucketing) -- it also blocks the execution harness, so it MUST
        # be answerable here via apply-kpi-panel-answer rather than leaving an
        # empty panel (the loop that burned operator quota). Advisory facets
        # remain set-only.
        if feature_questions:
            current = feature_questions[0]
        elif base_source_questions:
            current = base_source_questions[0]
        else:
            hard_intent = [
                q for q in intent_questions
                if str(q.get("facet")) in _HARD_BLOCKING_INTENT_FACETS
            ]
            current = (
                hard_intent[0]
                if hard_intent
                else _empty_panel(mapping, self.workspace, self.repo_root)
            )
        questions = feature_questions + base_source_questions + intent_questions
        # Attach the new "real-ops-dashboard" preview sections to every
        # generated panel so the renderer has data to work with.
        try:
            _attach_preview_sections(current, mapping, self.workspace, self.repo_root)
            for question in questions:
                if question is current:
                    continue
                _attach_preview_sections(question, mapping, self.workspace, self.repo_root)
        except Exception as exc:  # pragma: no cover - defensive
            # Preview composition must never break panel emission. If anything
            # in the executor / cache / redactor fails we drop the previews
            # silently; the rest of the panel still renders.
            current.setdefault("preview_compose_error", str(exc))

        # Conform to the canonical decision-panel contract (non-destructive: the
        # blocker panel keeps its own artifact_type, options, and evidence sections).
        from core.onboarding.panel_contract import normalize_decision_panel

        current.setdefault("stage", "blocker_question")
        normalize_decision_panel(current, workspace=self.workspace)

        current_json = self.output_dir / "current.json"
        current_markdown = self.output_dir / "current.md"
        current_full_markdown = self.output_dir / "current_full.md"
        index_json = self.output_dir / "index.json"
        current_json.write_text(json.dumps(current, indent=2, default=str) + "\n", encoding="utf-8")
        # current.md is the compact decision card -- it leads with the affected
        # KPI(s), the decision, and the options, which is all a human needs to
        # answer the blocker. The full evidence/proof render (feature-resolution
        # tables, source-truth packets, per-option proof packets, raw JSON) goes
        # to current_full.md for drill-down; current.json stays the machine
        # contract. Mirrors the results.md / results_full.md compact-vs-full
        # convention so operators are not buried under a wall of evidence.
        current_markdown.write_text(_render_markdown_compact(current), encoding="utf-8")
        current_full_markdown.write_text(_render_markdown(current), encoding="utf-8")
        # all_blockers.md: the whole open-decision surface at once (overview table
        # + each blocker's compact card), so an operator can SEE every pending
        # decision rather than only the single `current` one. Each is answerable
        # by feature via `apply-kpi-panel-answer --feature <name>`.
        (self.output_dir / "all_blockers.md").write_text(
            _render_all_blockers_overview(
                questions, _rel(self.workspace, self.repo_root)),
            encoding="utf-8",
        )
        index_json.write_text(
            json.dumps(
                {
                    "version": PANEL_VERSION,
                    "workspace": _rel(self.workspace, self.repo_root),
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                    "current_question": _rel(current_json, self.repo_root),
                    "questions": questions,
                },
                indent=2,
                default=str,
            )
            + "\n",
            encoding="utf-8",
        )

        return BlockerQuestionPanelResult(
            output_dir=_rel(self.output_dir, self.repo_root),
            current_json=_rel(current_json, self.repo_root),
            current_markdown=_rel(current_markdown, self.repo_root),
            current_full_markdown=_rel(current_full_markdown, self.repo_root),
            index_json=_rel(index_json, self.repo_root),
            question_count=len(questions),
            current_feature=str(current.get("feature", "")),
        )


def _deferred_kpi_ids_from_registry(workspace: Path) -> set[str]:
    """Derive the set of DEFERRED (undefined) KPI ids from the KPI registry.

    A KPI carries no measurable definition when BOTH its metric and grain/cuts
    are empty; such KPIs are deferred from this pass (not feature-blockers), so
    their unresolved feature tokens must not create blocker questions. Mirrors
    ``flow._undefined_kpis`` / the source-to-target planner / the validator so
    every gate agrees on one rule. The id is the registry entry's explicit
    ``kpi_id`` when present, else the generator's enumerated ``kpi_{idx:03d}``
    scheme. Generic (no domain vocabulary); never raises -> empty set on any
    missing/malformed registry so panel emission is never broken.
    """
    try:
        layout = WorkspaceLayout(project_root=workspace)
        registry_path = layout.contracts_dir / "kpi_registry.json"
        if not registry_path.exists():
            return set()
        registry = json.loads(registry_path.read_text(encoding="utf-8"))
    except Exception:
        return set()
    deferred: set[str] = set()
    for idx, kpi in enumerate(registry.get("kpis") or [], start=1):
        if not isinstance(kpi, dict):
            continue
        metric = str(kpi.get("metric") or "").strip()
        cuts = str(kpi.get("cuts") or "").strip()
        if metric or cuts:
            continue
        kpi_id = str(kpi.get("kpi_id") or "").strip() or f"kpi_{idx:03d}"
        deferred.add(kpi_id)
    return deferred


def _build_questions(
    mapping: dict[str, Any],
    workspace: Path,
    repo_root: Path,
    deferred_kpi_ids: set[str] | None = None,
) -> list[dict[str, Any]]:
    feature_items = _feature_items(mapping, deferred_kpi_ids)
    clusters = mapping.get("blocker_clusters") or []
    if not clusters:
        clusters = _clusters_from_features(feature_items)
    else:
        # Drop any pre-computed cluster whose unresolved items all belong to
        # deferred KPIs -- nothing left to ask once those features are filtered.
        clusters = [
            cluster
            for cluster in clusters
            if feature_items.get(_norm(str(cluster.get("feature") or "")))
        ]
    questions = []
    for cluster in clusters:
        feature = str(cluster.get("feature") or "")
        if not feature:
            continue
        items = feature_items.get(_norm(feature), [])
        if not items:
            continue
        questions.append(_question_for_cluster(mapping, workspace, repo_root, cluster, items))
    return questions


def _feature_items(
    mapping: dict[str, Any],
    deferred_kpi_ids: set[str] | None = None,
) -> dict[str, list[dict[str, Any]]]:
    deferred = set(deferred_kpi_ids or set())
    items: dict[str, list[dict[str, Any]]] = {}
    for kpi in mapping.get("kpis", []):
        # Partial-completion: skip deferred (undefined) KPIs so their unresolved
        # feature tokens do not become feature-blocker questions. Defined KPIs
        # still surface their own blockers normally. Exception: a
        # `no_supporting_evidence` blocker is about the KPI itself (its prose
        # matched nothing in the workspace), not a sibling feature token — the
        # user must be asked to confirm the missing data or point at the
        # source, so it survives deferral.
        kpi_deferred = str(kpi.get("kpi_id") or "") in deferred
        for feature in kpi.get("features", []):
            if feature.get("state") in READY_STATES:
                continue
            if kpi_deferred and feature.get("resolution_type") != "no_supporting_evidence":
                continue
            name = str(feature.get("feature") or "")
            items.setdefault(_norm(name), []).append({"kpi": kpi, "feature": feature})
    return items


def _clusters_from_features(items: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    clusters = []
    for feature_items in items.values():
        first = feature_items[0]["feature"]
        clusters.append(
            {
                "feature": first.get("feature", ""),
                "count": len(feature_items),
                "risk": "unknown",
                "examples": [item["kpi"].get("kpi_id", "") for item in feature_items[:3]],
            }
        )
    clusters.sort(key=lambda item: (-int(item.get("count", 0)), str(item.get("feature", ""))))
    return clusters


def _question_for_cluster(
    mapping: dict[str, Any],
    workspace: Path,
    repo_root: Path,
    cluster: dict[str, Any],
    items: list[dict[str, Any]],
) -> dict[str, Any]:
    feature = str(cluster.get("feature") or items[0]["feature"].get("feature") or "")
    derived_options = _valid_derived_options(items)
    physical_options = _physical_column_options(items)
    applies_to = [str(item["kpi"].get("kpi_id") or "") for item in items]
    evidence_files = _evidence_files(items, repo_root)
    source_truth = _kpi_source_truth(items, workspace, repo_root)
    kpi_understanding = _kpi_understanding_packet(items, source_truth, feature)
    if not physical_options:
        physical_options = _profile_candidate_options(items, workspace, repo_root, feature)
    base = {
        "artifact_type": "blocker_question_panel/current.json",
        "version": PANEL_VERSION,
        "generated_by": "blocker-question-panel",
        "workspace": _rel(workspace, repo_root),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "question_id": f"workspace_{_slug(feature)}",
        "feature": feature,
        "applies_to_kpis": applies_to,
        "reuse_scope": "workspace_level" if len(applies_to) > 1 else "kpi_specific",
        "risk": cluster.get("risk", "unknown"),
        "status": "needs_user_answer",
        "source_mapping": _rel(workspace / "interns" / "generated" / "contracts" / "kpi_feature_mapping.json", repo_root),
        "evidence_files": evidence_files,
        "interaction_contract": INTERACTION_CONTRACT,
        "panel_contract": {
            "display_shape": "full_kpi_truth_packet",
            "source_truth_required": True,
            "option_proof_required": True,
            "default_code_preference": "sql",
            "must_include": [
                "absolute KPI source truth",
                "all KPI fields",
                "source cell or source artifact proof",
                "recommended option",
                "formula or derived logic when present",
                "SQL table/column mapping with source evidence",
                "sample values",
                "SQL query or query sketch",
                "demo result table shape",
                "custom fallback option",
            ],
        },
        "default_code_preference": "sql",
        "output_dialect": {
            "default": "sql",
            "label": "SQL (default)",
            "alternatives": ["polars", "pyspark", "databricks_sql"],
            "rule": "Render SQL by default. Generate other dialects only when the user explicitly selects them.",
        },
        "immutable_kpi_policy": {
            "rule": "The KPI from the source workbook or registry is hard truth and must not be rewritten.",
            "understanding_is_review_context_only": True,
            "placeholder_sql_is_non_executable": True,
        },
        "kpi_source_truth": source_truth,
        "kpi_understanding": kpi_understanding,
    }
    prior = _prior_wiki_decision(workspace, repo_root, feature)
    if prior:
        base["prior_decision_wiki"] = prior
    no_evidence_items = [
        item
        for item in items
        if item["feature"].get("resolution_type") == "no_supporting_evidence"
    ]
    if no_evidence_items:
        kpi_labels = ", ".join(
            str(item["kpi"].get("kpi_id") or "") for item in no_evidence_items
        )
        feature_question = next(
            (
                str(item["feature"].get("question") or "")
                for item in no_evidence_items
                if item["feature"].get("question")
            ),
            "",
        )
        return {
            **base,
            "blocker_label": "no_supporting_evidence",
            "blocker": (
                f"No workspace evidence supports KPI(s) {kpi_labels}: no term in "
                "the KPI prose matches any profiled column, dataset name, "
                "data-dictionary entry, or accepted workspace definition. The "
                "KPI may presuppose data this workspace does not contain."
            ),
            "question": feature_question
            or (
                "Does this KPI presuppose data that does not exist in this "
                "workspace? Confirm that, or point to the source dataset, "
                "column, or file that holds the data."
            ),
            "answer_type": "no_supporting_evidence",
            # option_a/option_b carry expected_answer_shape placeholders the
            # user must fill in, so neither is appliable as-recommended;
            # `custom` is the only recommendation the validator accepts here.
            "recommended_option_id": "custom",
            "recommended_answer": (
                "Confirm the workspace does not contain the presupposed data, "
                "or point to the source that holds it. No mapping will be "
                "fabricated from zero evidence."
            ),
            "why": (
                "Every prose term of this KPI was scanned against the "
                "workspace's profiled columns, dataset names, dictionary "
                "descriptions, and accepted definitions, and nothing matched. "
                "Mapping it anyway would fabricate evidence; the honest options "
                "are to confirm the data is absent or to supply the missing "
                "source."
            ),
            "options": [
                {
                    "option_id": "option_a",
                    "label": "Confirm the data does not exist in this workspace",
                    "business_summary": (
                        "Record that this KPI presupposes data the workspace "
                        "does not contain. The KPI stays blocked as not "
                        "computable from this workspace until a source is "
                        "supplied."
                    ),
                    "expected_answer_shape": {
                        "confirmation": "data_not_in_workspace",
                        "reason": "",
                        "confirmed_by": "",
                        "applies_to_kpis": applies_to,
                    },
                    "json_backed": False,
                },
                {
                    "option_id": "option_b",
                    "label": "Point to the source that holds this data",
                    "business_summary": (
                        "Name the dataset, column(s), file, or upstream system "
                        "that contains the data this KPI needs, so it can be "
                        "onboarded and profiled as evidence."
                    ),
                    "expected_answer_shape": {
                        "source_dataset_or_file": "",
                        "columns": [],
                        "owner_or_system": "",
                        "reason": "",
                        "applies_to_kpis": applies_to,
                    },
                    "json_backed": False,
                },
                _custom_rule_option(feature),
            ],
        }
    conflict_items = [
        item
        for item in items
        if item["feature"].get("resolution_type") == "dictionary_conflict"
    ]
    if conflict_items:
        return _dictionary_conflict_question(base, feature, applies_to, conflict_items)
    if any(item["feature"].get("resolution_type") == "kpi_definition_required" for item in items):
        definition_items = [
            item
            for item in items
            if item["feature"].get("resolution_type") == "kpi_definition_required"
        ]
        has_prose = any(
            isinstance(evidence, dict) and evidence.get("type") == "kpi_prose"
            for item in definition_items
            for evidence in item["feature"].get("evidence") or []
        )
        feature_question = next(
            (
                str(item["feature"].get("question") or "")
                for item in definition_items
                if item["feature"].get("question")
            ),
            "",
        )
        if has_prose:
            blocker_text = (
                "The KPI is defined in stakeholder prose; a concrete metric "
                "expression and grain have not been confirmed yet. The prose and "
                "matched workspace evidence are attached."
            )
            question_text = feature_question or (
                "Which concrete metric expression and grain/dimensions implement "
                "this KPI's prose definition?"
            )
        else:
            blocker_text = (
                "The current KPI registry contains a seed placeholder KPI, not a concrete "
                "business metric that can be mapped to data."
            )
            question_text = (
                "Which concrete KPI should replace the seed? Include the business question, "
                "metric expression, grain/dimensions, owner, and acceptance tests."
            )
        return {
            **base,
            "blocker": blocker_text,
            "question": question_text,
            "answer_type": "kpi_definition_required",
            "recommended_option_id": "custom",
            "recommended_answer": "Provide a concrete KPI definition before mapping features.",
            "why": (
                "Executable KPI logic needs a proven metric and grain. Mapping placeholder words "
                "such as confirm, metric, or grain to columns would create invalid evidence."
            ),
            "options": [
                {
                    "option_id": "option_a",
                    "label": "Provide KPI definition",
                    "business_summary": (
                        "Replace the seed KPI with a concrete metric — a rate, trend, "
                        "aging, or count — with a defined business question and grain."
                    ),
                    "expected_answer_shape": {
                        "business_question": "",
                        "metric": "",
                        "grain_or_cuts": "",
                        "owner": "",
                        "acceptance_tests": [],
                        "evidence_source": "",
                        "applies_to_kpis": applies_to,
                    },
                    "json_backed": False,
                },
                {
                    "option_id": "custom",
                    "label": "Restart KPI generation",
                    "business_summary": (
                        "Run KPI generation again with stakeholder context or a richer registry source."
                    ),
                    "expected_answer_shape": {
                        "context_file": "",
                        "generation_notes": "",
                        "applies_to_kpis": applies_to,
                    },
                    "json_backed": False,
                },
            ],
        }
    if derived_options:
        return {
            **base,
            "blocker": (
                f"`{feature}` is unresolved and has JSON-backed candidate derivation "
                f"option(s) for {len(applies_to)} KPI(s)."
            ),
            "question": f"Which definition should be accepted for `{feature}`?",
            "answer_type": "select_json_backed_option_or_custom_rule",
            "recommended_option_id": "custom",
            "recommended_answer": (
                "Provide or confirm an authoritative business definition; accept a JSON-backed "
                "candidate only after confirming the formula, source, and grain."
            ),
            "why": (
                "JSON-backed derived options include formula, inputs, observed values, evidence sources, "
                "and derivation reasoning, but they remain candidate evidence rather than ground truth."
            ),
            "options": [
                _derived_option_payload(option, idx, source_truth)
                for idx, option in enumerate(derived_options, start=1)
            ]
            + [_custom_rule_option(feature)],
        }
    if physical_options:
        top_options = physical_options[:3]
        overflow_count = max(0, len(physical_options) - len(top_options))
        top_choice = top_options[0]
        top_reason = str(top_choice.get("reason") or "highest profile-evidence score")
        top_score = float(top_options[0].get("score") or 0) if top_options else 0.0
        second_score = float(top_options[1].get("score") or 0) if len(top_options) > 1 else 0.0
        # A "recommended" label must survive the same discipline as an auto-proven
        # resolver match: a real bar (not merely being first in a list) and a real
        # margin over the runner-up. Mirrors feature_resolver.py's own auto-proven
        # check in spirit, scaled to this function's own score range.
        recommend_top = top_score >= 60 and (len(top_options) == 1 or top_score - second_score >= 20)
        kpi_list = ", ".join(str(k) for k in applies_to[:3])
        kpi_suffix = f" (+{len(applies_to)-3} more)" if len(applies_to) > 3 else ""
        return {
            **base,
            "preamble": (
                f"Resolving `{feature}` for KPI(s): {kpi_list}{kpi_suffix}. "
                "The platform scanned every profiled column and ranked candidates "
                "by name match, KPI-text overlap, and dataset-name overlap. "
                f"Showing top {len(top_options)} of {len(physical_options)} candidates."
            ),
            "blocker": (
                f"`{feature}` is unresolved for {len(applies_to)} KPI(s), and multiple "
                "profile-backed physical column candidates are available."
            ),
            "question": f"Which physical column should define `{feature}` as a workspace-level mapping?",
            "answer_type": "select_physical_column_or_custom_rule",
            "recommended_option_id": "option_a",
            "recommended_answer": (
                f"Accept Option A — `{_sql_column_label(str(top_choice.get('dataset') or ''), str(top_choice.get('column') or ''))}` — "
                f"because: {top_reason}."
            ),
            "why": (
                "These options come from schema/profile alias evidence. They are candidate mappings, "
                "not accepted business truth until confirmed."
            ),
            "overflow_options_count": overflow_count,
            "overflow_options_pointer": (
                f"+ {overflow_count} more lower-scoring options in `current.json` under "
                "`hidden_overflow_options`. Pass --answer with the explicit option_id to pick one."
                if overflow_count else ""
            ),
            "options": [
                _physical_option_payload(option, idx, source_truth, is_recommended=(idx == 1 and recommend_top))
                for idx, option in enumerate(top_options, start=1)
            ]
            + [_custom_rule_option(feature)],
            "hidden_overflow_options": [
                _physical_option_payload(option, idx + len(top_options), source_truth, is_recommended=False)
                for idx, option in enumerate(physical_options[len(top_options):], start=1)
            ],
        }
    evidence_pack = _cli_agent_evidence_pack(feature, items, workspace, repo_root)
    if evidence_pack["available_columns"]:
        return {
            **base,
            "blocker": (
                f"`{feature}` is unresolved for {len(applies_to)} KPI(s). No evidence-backed "
                "derived or physical option could be produced from the workspace alone, but "
                f"{len(evidence_pack['available_columns'])} profile-scanned column(s) are available "
                "as bounded evidence for a semantic proposal."
            ),
            "question": (
                f"Propose a mapping for `{feature}` from the bounded evidence below, then ask the "
                "user to confirm before applying."
            ),
            "answer_type": "cli_agent_proposal_needed",
            "recommended_option_id": "option_a",
            "recommended_answer": (
                "The orchestrating CLI agent should read `cli_agent_evidence_pack`, propose a JSON "
                "mapping in the shape of `option_a.expected_answer_shape`, present it to the user, "
                "and only after explicit user approval call apply-kpi-panel-answer with "
                "--answer custom --custom-definition <agent-proposed JSON>."
            ),
            "why": (
                "The Python resolver does not call any LLM directly. When workspace evidence alone "
                "cannot resolve a feature, the orchestrating CLI agent (whichever LLM-backed CLI is "
                "currently running) proposes a mapping from the bounded evidence pack. The proposal "
                "is non-executable until the user confirms it via the normal apply path."
            ),
            "cli_agent_evidence_pack": evidence_pack,
            "cli_agent_task": _cli_agent_task_text(feature, applies_to, evidence_pack, workspace, repo_root),
            "options": [
                _cli_agent_proposal_option(feature, evidence_pack),
                _custom_rule_option(feature),
            ],
        }
    return {
        **base,
        "blocker": (
            f"`{feature}` is unresolved for {len(applies_to)} KPI(s), and no valid JSON-backed "
            "derived formula option is currently available."
        ),
        "question": f"What authoritative source, physical column, or accepted workspace rule defines `{feature}`?",
        "answer_type": "direct_mapping_or_business_rule",
        "recommended_option_id": "custom",
        "recommended_answer": "Provide a concrete source-backed mapping, formula, or business rule.",
        "why": (
            "A formula should not be invented when the resolver cannot produce a valid evidence-backed "
            "derived option. This answer can be saved as a reusable workspace definition."
        ),
        "options": [
            _custom_rule_option(feature),
        ],
    }


def _dictionary_conflict_question(
    base: dict[str, Any],
    feature: str,
    applies_to: list[str],
    conflict_items: list[dict[str, Any]],
) -> dict[str, Any]:
    """Panel question for a feature whose dictionary claim contradicts profiles.

    The conflicted column was documented one way and observed another, so the
    user must rule which evidence wins. option_a/option_b carry
    expected_answer_shape placeholders the user must fill in, so neither is
    appliable as-recommended; `custom` is the only recommendation the
    validator accepts here (same rule as no_supporting_evidence).
    """
    conflicts: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in conflict_items:
        for conflict in item["feature"].get("conflicts") or []:
            if not isinstance(conflict, dict):
                continue
            conflict_id = str(conflict.get("conflict_id") or "")
            if conflict_id and conflict_id in seen:
                continue
            seen.add(conflict_id)
            conflicts.append(conflict)
    errors = [c for c in conflicts if c.get("severity") == "error"]
    lead = errors[0] if errors else (conflicts[0] if conflicts else {})
    dataset_name = str(lead.get("dataset_name") or "")
    column = str(lead.get("column") or feature)
    column_label = f"{dataset_display_stem(dataset_name) or 'dataset'}.{column}"
    feature_question = next(
        (
            str(item["feature"].get("question") or "")
            for item in conflict_items
            if item["feature"].get("question")
        ),
        "",
    )
    return {
        **base,
        "blocker_label": "dictionary_conflict",
        "blocker": (
            f"`{feature}` resolves to `{column_label}`, but the data "
            "dictionary's documented claim about that column contradicts "
            "profiled evidence. Documentation and data cannot both be right, "
            "so the mapping is blocked until a human rules which wins."
        ),
        "question": feature_question
        or (
            f"The dictionary's claim about `{column_label}` conflicts with "
            "profiled evidence. Should the observed data be trusted over the "
            "dictionary (and how should the column be interpreted), or is the "
            "dictionary right and the data needs normalization first?"
        ),
        "answer_type": "dictionary_conflict",
        "recommended_option_id": "custom",
        "recommended_answer": (
            "Rule which evidence wins for this column: trust the observed "
            "data (state the interpretation to use), trust the dictionary "
            "(state the normalization/repair required first), or provide a "
            "different source. The documented claim will not be silently "
            "trusted over contradicting profile evidence."
        ),
        "why": (
            "The data dictionary is documentation evidence, not ground truth. "
            "Profile evidence actively contradicts its claim for this column "
            "(see `dictionary_conflicts`), so using the column as documented "
            "could silently produce wrong KPI numbers."
        ),
        "dictionary_conflicts": conflicts,
        "options": [
            {
                "option_id": "option_a",
                "label": "Trust the observed data over the dictionary",
                "business_summary": (
                    "Treat the dictionary entry as stale or wrong. Provide the "
                    "interpretation the profiled values support (e.g. the real "
                    "code meanings, the per-row unit rule, or the correct "
                    "source column) and record it as a workspace definition."
                ),
                "expected_answer_shape": {
                    "ruling": "trust_observed_data",
                    "feature": feature,
                    "interpretation": "",
                    "source_columns": [],
                    "reason": "",
                    "confirmed_by": "",
                    "applies_to_kpis": applies_to,
                },
                "json_backed": False,
            },
            {
                "option_id": "option_b",
                "label": "The dictionary is right; the data needs repair first",
                "business_summary": (
                    "Treat the profiled values as bad data. State the "
                    "normalization or upstream fix required before this column "
                    "may feed any KPI; the KPI stays blocked until then."
                ),
                "expected_answer_shape": {
                    "ruling": "trust_dictionary_claim",
                    "feature": feature,
                    "required_data_fix": "",
                    "reason": "",
                    "confirmed_by": "",
                    "applies_to_kpis": applies_to,
                },
                "json_backed": False,
            },
            _custom_rule_option(feature),
        ],
    }


def _prior_wiki_decision(workspace: Path, repo_root: Path, feature: str) -> dict[str, Any] | None:
    note = read_feature_note(WikiLayout(project_root=workspace), feature)
    if note is None:
        return None
    fm = note.frontmatter or {}
    return {
        "path": _rel(note.path, repo_root),
        "summary": fm.get("summary") or "",
        "updated": fm.get("updated") or "",
        "user_why": note.why,
        "has_user_why": note.has_user_why,
    }


def _kpi_source_truth(items: list[dict[str, Any]], workspace: Path, repo_root: Path) -> list[dict[str, Any]]:
    registry = _load_json(workspace / "interns" / "generated" / "contracts" / "kpi_registry.json")
    registry_kpis = registry.get("kpis") if isinstance(registry.get("kpis"), list) else []
    registry_by_id = {
        f"kpi_{idx:03d}": kpi
        for idx, kpi in enumerate(registry_kpis, start=1)
        if isinstance(kpi, dict)
    }
    rows = []
    seen = set()
    for item in items:
        kpi = item["kpi"]
        kpi_id = str(kpi.get("kpi_id") or "")
        if kpi_id in seen:
            continue
        seen.add(kpi_id)
        source = kpi | {key: value for key, value in registry_by_id.get(kpi_id, {}).items() if value not in (None, "")}
        source_path = str(source.get("source") or "")
        cell_trace = _excel_cell_trace(source_path, source) if source_path else {}
        rows.append(
            {
                "kpi_id": kpi_id,
                "business_question": str(source.get("name") or ""),
                "description": str(source.get("description") or ""),
                "metric": str(source.get("metric") or ""),
                "cuts": _split_cuts(str(source.get("cuts") or "")),
                "source": _source_label(source_path, repo_root),
                "source_cells": cell_trace.get("cells", {}),
                "source_sheet": cell_trace.get("sheet", ""),
                "source_truth_note": (
                    "Source workbook text is authoritative for KPI question, description, "
                    "metric wording, cuts, filters, and continuation rows. It is not proof of "
                    "joins, derived formulas, or executable grain unless those are explicit."
                ),
            }
        )
    return rows


def _kpi_understanding_packet(
    items: list[dict[str, Any]],
    source_truth: list[dict[str, Any]],
    feature: str,
) -> list[dict[str, Any]]:
    truth_by_id = {str(item.get("kpi_id") or ""): item for item in source_truth}
    rows = []
    seen = set()
    for item in items:
        kpi = item["kpi"]
        kpi_id = str(kpi.get("kpi_id") or "")
        if kpi_id in seen:
            continue
        seen.add(kpi_id)
        truth = truth_by_id.get(kpi_id, {})
        metric = str(truth.get("metric") or kpi.get("metric") or "")
        cuts = list(truth.get("cuts") or _split_cuts(str(kpi.get("cuts") or "")))
        question = str(truth.get("business_question") or kpi.get("name") or "")
        semantic = _is_semantic_blocker(feature, metric, cuts, question, kpi)
        rows.append(
            {
                "kpi_id": kpi_id,
                "presentation_level": "full" if semantic else "compact",
                "requires_understanding_approval": semantic,
                "affected_unresolved_feature": feature,
                "original_kpi": {
                    "business_question": question,
                    "description": str(truth.get("description") or kpi.get("description") or ""),
                    "metric": metric,
                    "cuts": cuts,
                    "source": str(truth.get("source") or kpi.get("source") or ""),
                },
                "my_understanding": _understanding_text(question, metric, cuts),
                "understanding_warning": (
                    "This is interpretation for review only. It does not replace or modify the source KPI."
                ),
                "output_dialect": "SQL",
                "strict_proven_sql": _strict_proven_sql(kpi),
                "intent_sql_sketch": _intent_sql_sketch(kpi, feature, metric, cuts),
                "demo_result_table": _kpi_demo_table(kpi_id, metric, cuts),
            }
        )
    return rows


def _is_semantic_blocker(
    feature: str,
    metric: str,
    cuts: list[str],
    question: str,
    kpi: dict[str, Any],
) -> bool:
    haystack = " ".join([feature, metric, question, ", ".join(cuts)]).lower()
    # Generic semantic tokens — math/SQL/temporal vocabulary common to every
    # business domain. Domain-specific filter terms (payer, LOB, etc.) used
    # to live here; they now come from `workspace_vocabulary.json` filter_terms.
    semantic_tokens = {
        "percentage", "share", "denominator", "grain",
        "age", "date", "month", "quarter", "year",
        "top", "join", "distinct",
        "sum(", "avg(", "count(", "min(", "max(",
    }
    if any(token in haystack for token in semantic_tokens):
        return True
    # Pick up workspace-derived filter literals as additional semantic tokens
    # (e.g., "Medicare" for healthcare, "Wholesale" for retail).
    try:
        from core.onboarding.lexicon.vocabulary import terms_for as _vt
        from core.storage.workspace_layout import WorkspaceLayout
        from core.paths import PROJECT_ROOT as _ROOT
        workspace_path = kpi.get("workspace") or ""
        if workspace_path:
            layout = WorkspaceLayout(project_root=(_ROOT / str(workspace_path)).resolve())
            for term in _vt(layout, "filter_terms"):
                if str(term).lower() in haystack:
                    return True
    except Exception:
        pass
    return any(
        feature_item.get("resolution_type") in {"derived_formula", "kpi_definition_required"}
        for feature_item in kpi.get("features", [])
    )


def _understanding_text(question: str, metric: str, cuts: list[str]) -> str:
    parts = [f"Answer the source KPI exactly as written: {question}"]
    if metric:
        parts.append(f"Compute `{metric}`.")
    if cuts:
        parts.append("Break out or filter by: " + ", ".join(cuts) + ".")
    return " ".join(parts)


def _strict_proven_sql(kpi: dict[str, Any]) -> str:
    ready_features = [
        feature
        for feature in kpi.get("features", [])
        if feature.get("state") in READY_STATES and feature.get("source_columns")
    ]
    if not ready_features:
        return "-- No strict proven SQL yet: this KPI has no fully proven source mappings for the current preview."
    select_items = []
    table = "proven_source"
    for feature in ready_features[:8]:
        source = (feature.get("source_columns") or [{}])[0]
        table = dataset_display_stem(str(source.get("dataset") or "")) or table
        column = str(source.get("column") or "")
        if column:
            select_items.append(f'  "{column}" AS "{_slug(str(feature.get("feature") or column))}"')
    if not select_items:
        return "-- No strict proven SQL yet: proven mappings have no physical columns."
    return "SELECT\n" + ",\n".join(select_items) + f'\nFROM "{table}"\nLIMIT 20;'


def _intent_sql_sketch(
    kpi: dict[str, Any],
    feature: str,
    metric: str,
    cuts: list[str],
) -> str:
    metric_expr = metric or "<METRIC_EXPRESSION>"
    table = _first_kpi_table(kpi) or "<SOURCE_TABLE>"
    group_columns = [
        _placeholder_for_cut(cut)
        for cut in cuts
        if not any(token in cut for token in ("=", ">", "<"))
    ][:6]
    filters = [cut for cut in cuts if any(token in cut for token in ("=", ">", "<"))]
    select_lines = [f"  {metric_expr} AS metric_value"]
    select_lines.extend(f"  {column}" for column in group_columns)
    lines = [
        "-- NON-EXECUTABLE INTENT SKETCH: placeholders require user/proof confirmation.",
        "-- KPI text is hard truth; this sketch is only to review intent.",
        "SELECT",
        ",\n".join(select_lines),
        f"FROM {table}",
    ]
    if filters:
        lines.extend(["WHERE " + " AND ".join(filters)])
    if group_columns:
        lines.append("GROUP BY " + ", ".join(group_columns))
    lines.append(f"-- unresolved blocker: <{feature}>")
    return "\n".join(lines) + ";"


def _first_kpi_table(kpi: dict[str, Any]) -> str:
    for feature in kpi.get("features", []):
        for source in feature.get("source_columns") or []:
            dataset = str(source.get("dataset") or "")
            if dataset:
                return '"' + (dataset_display_stem(dataset) or "source_table") + '"'
    return ""


def _placeholder_for_cut(cut: str) -> str:
    clean = re.sub(r"\([^)]*\)", "", cut)
    clean = re.sub(r"[^A-Za-z0-9_]+", "_", clean).strip("_")
    return f"<{clean or 'DIMENSION'}>"


def _kpi_demo_table(kpi_id: str, metric: str, cuts: list[str]) -> str:
    dimensions = [
        _placeholder_for_cut(cut).strip("<>")
        for cut in cuts
        if not any(token in cut for token in ("=", ">", "<"))
    ][:3]
    row: dict[str, Any] = {dimension: "<example>" for dimension in dimensions}
    metric_name = _slug(metric or "metric_value")
    row[metric_name] = "<computed>"
    row["kpi_id"] = kpi_id
    return _markdown_table([row])


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return data if isinstance(data, dict) else {}


def _split_cuts(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def _source_label(path: str, repo_root: Path) -> str:
    if not path:
        return ""
    source_path = Path(path)
    return _rel(source_path, repo_root) if source_path.is_absolute() else path


def _sql_table_label(path: str) -> str:
    return dataset_display_stem(str(path or "")) or "source_table"


def _sql_column_label(path: str, column: str) -> str:
    return f"{_sql_table_label(path)}.{column or 'unknown'}"


def _excel_cell_trace(source_path: str, source: dict[str, Any]) -> dict[str, Any]:
    path = Path(source_path)
    if not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        wb = load_workbook(path, data_only=False, read_only=False)
    except Exception:
        return {}
    target_question = str(source.get("name") or "").strip()
    if not target_question:
        return {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            first = row[0].value if row else None
            if str(first or "").strip() != target_question:
                continue
            row_number = row[0].row
            cells = {
                "business_question": row[0].coordinate,
                "description": row[1].coordinate if len(row) > 1 else "",
                "first_cut": row[2].coordinate if len(row) > 2 else "",
                "metric": row[3].coordinate if len(row) > 3 else "",
            }
            cut_cells = []
            for scan_row in range(row_number, ws.max_row + 1):
                next_question = ws.cell(scan_row, 1).value
                if scan_row != row_number and next_question not in (None, ""):
                    break
                cut_value = ws.cell(scan_row, 3).value
                if cut_value not in (None, ""):
                    cut_cells.append(ws.cell(scan_row, 3).coordinate)
            if cut_cells:
                cells["cuts"] = ":".join([cut_cells[0], cut_cells[-1]]) if len(cut_cells) > 1 else cut_cells[0]
            return {"sheet": ws.title, "cells": cells}
    return {}


def _valid_derived_options(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    options = []
    seen = set()
    for item in items:
        for option in item["feature"].get("derived_feature_options") or []:
            if not option.get("input_columns"):
                continue
            if option.get("missing_inputs"):
                continue
            key = (
                option.get("derived_column_name"),
                option.get("formula"),
                option.get("source_pattern_id"),
            )
            if key in seen:
                continue
            seen.add(key)
            options.append(option)
    return options


def _physical_column_options(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    options = []
    seen = set()
    for item in items:
        feature = item["feature"]
        for column in feature.get("source_columns") or []:
            key = (column.get("dataset"), column.get("column"))
            if key in seen:
                continue
            seen.add(key)
            options.append(
                {
                    "feature_state": feature.get("state"),
                    "resolution_type": feature.get("resolution_type"),
                    "kpi_id": item["kpi"].get("kpi_id"),
                    "kpi_name": item["kpi"].get("name"),
                    "kpi_metric": item["kpi"].get("metric"),
                    "dataset": column.get("dataset"),
                    "column": column.get("column"),
                    "dtype": column.get("dtype"),
                    "row_count": column.get("row_count"),
                    "score": column.get("score"),
                    "profile_path": column.get("profile_path"),
                    "observed_values": column.get("observed_values") or [],
                    "value_profile": column.get("value_profile") or {},
                    "semantic_meaning_sources": column.get("semantic_meaning_sources") or [],
                    "mapping_proof": column.get("mapping_proof") or {},
                    "answer_demo": _answer_demo(item["kpi"], feature, column),
                    "evidence_state": "schema_profile_alias_candidate",
                    "reason": column.get("reason")
                    or "Column matched the unresolved KPI term through configured business/schema aliases.",
                }
            )
    # See `_profile_candidate_options` below for the full story: a relation this
    # platform wrote is never a better source of truth than a base table, and a
    # score tie must not be broken ALPHABETICALLY -- that is how a KPI's own
    # `*_results` view came to be the recommended definition of that KPI's cut.
    from core.onboarding.workspace.onboarding import _is_platform_written_relation

    options.sort(
        key=lambda item: (
            _is_platform_written_relation(str(item.get("dataset") or "")),
            -float(item.get("score") or 0),
            str(item.get("dataset") or ""),
            str(item.get("column") or ""),
        )
    )
    return options


def _profile_candidate_options(
    items: list[dict[str, Any]],
    workspace: Path,
    repo_root: Path,
    feature: str,
) -> list[dict[str, Any]]:
    if _norm(feature) in {"distinct", "disitnct"}:
        return []
    profile_index = _load_json(workspace / "interns" / "generated" / "profiles" / "profile_index.json")
    profiles = profile_index.get("profiles") if isinstance(profile_index.get("profiles"), list) else []
    scored: list[dict[str, Any]] = []
    for profile in profiles:
        dataset = str(profile.get("path") or "")
        if not dataset:
            continue
        try:
            if not WorkspaceLayout(workspace).is_dataset_allowed(Path(dataset)):
                continue
        except OSError:
            continue
        for column in profile.get("columns") or []:
            column_name = str(column.get("name") or "")
            score, reason = _profile_candidate_score(feature, dataset, column_name, items)
            if score <= 0:
                continue
            scored.append(
                {
                    "feature_state": "profile_candidate",
                    "resolution_type": "profile_inferred_physical_column",
                    "kpi_id": items[0]["kpi"].get("kpi_id"),
                    "kpi_name": items[0]["kpi"].get("name"),
                    "kpi_metric": items[0]["kpi"].get("metric"),
                    "dataset": _rel(Path(dataset), repo_root),
                    "column": column_name,
                    "dtype": column.get("dtype"),
                    "row_count": profile.get("row_count"),
                    "score": score,
                    "profile_path": profile.get("profile_path"),
                    "observed_values": _sample_values(column.get("sample_values")),
                    "value_profile": {
                        "sample_values": _sample_values(column.get("sample_values")),
                        "sample_min": column.get("sample_min"),
                        "sample_max": column.get("sample_max"),
                        "null_count": column.get("null_count"),
                        "source": column.get("source") or "sample_profile",
                    },
                    "semantic_meaning_sources": [
                        {
                            "source": profile.get("profile_path"),
                            "evidence_type": "profile_index_column_sample",
                            "reason": reason,
                        }
                    ],
                    "mapping_proof": {
                        "sample_query": (
                            f'SELECT "{column_name}" AS "{_slug(column_name)}" '
                            f'FROM "{dataset_display_stem(dataset)}" LIMIT 5;'
                        ),
                        "sample_output": [
                            {_slug(column_name): value}
                            for value in _sample_values(column.get("sample_values"))[:5]
                        ],
                        "source_files": [profile.get("profile_path")],
                    },
                    "evidence_state": "profile_scanned_candidate",
                    "reason": reason,
                }
            )
    # A relation THIS PLATFORM wrote is never a better source of truth than a
    # base table, at any score. Observed 2026-07-27: a KPI's own `*_results`
    # view tied on score with the real dimension table and won the ALPHABETICAL
    # tiebreak, so the panel recommended defining that KPI's cut from that same
    # KPI's output. Profiling should no longer surface such relations at all
    # (see `_is_platform_written_relation`), but a profile index written before
    # that fix still carries them, and a ranking that can prefer a derived
    # relation is wrong on its own terms.
    from core.onboarding.workspace.onboarding import _is_platform_written_relation

    scored.sort(
        key=lambda item: (
            _is_platform_written_relation(str(item.get("dataset") or "")),
            -float(item.get("score") or 0),
            str(item.get("dataset") or ""),
            str(item.get("column") or ""),
        )
    )
    return scored[:20]


def _profile_candidate_score(
    feature: str,
    dataset: str,
    column: str,
    items: list[dict[str, Any]],
) -> tuple[float, str]:
    feature_norm = _norm(feature)
    column_norm = _norm(column)
    dataset_norm = _norm(dataset_display_stem(dataset))
    kpi_text = _norm(
        " ".join(
            str(value or "")
            for item in items
            for value in [item["kpi"].get("name"), item["kpi"].get("description"), item["kpi"].get("cuts")]
        )
    )
    score = 0.0
    reasons: list[str] = []
    if feature_norm and feature_norm == column_norm:
        score += 100
        reasons.append("column name exactly matches unresolved feature")
    elif feature_norm and (feature_norm in column_norm or column_norm in feature_norm):
        score += 60
        reasons.append("column name partially matches unresolved feature")

    # Direct feature-dataset alignment: table named after the feature is the
    # strongest non-lexical signal that this dataset is the right one.
    if feature_norm and dataset_norm and feature_norm == dataset_norm.rstrip("s"):
        score += 30
        reasons.append("dataset name directly aligns with feature term")

    # Workspace-derived aliases live in the workspace lexicon; this scorer
    # stays domain-agnostic. The previous _feature_synonyms() dict (which
    # hardcoded "department"/"lob" healthcare-RCM vocabulary) has been removed.
    if column_norm and column_norm in kpi_text:
        score += 20
        reasons.append("column name appears in KPI text")
    if dataset_norm and dataset_norm.rstrip("s") in kpi_text:
        score += 20
        reasons.append("dataset name appears in KPI text")
    if column_norm in {"id", "insertdate", "modifieddate"}:
        score -= 30
        reasons.append("generic technical column")
    # Penalise columns that are clearly surrogate/foreign keys (end in "id",
    # e.g. "patientid", "claimid") when the feature name doesn't appear in
    # the column name — they identify records but don't describe the feature.
    elif column_norm.endswith("id") and len(column_norm) > 2:
        if not (feature_norm and feature_norm in column_norm):
            score -= 30
            reasons.append("column is a key/ID column not matching the feature")
    return score, "; ".join(reasons) or "profile column candidate"


def _sample_values(raw: Any) -> list[Any]:
    if raw is None:
        return []
    if isinstance(raw, list):
        return raw[:8]
    text = str(raw).strip()
    if not text:
        return []
    return text.split()[:8]


def _physical_option_payload(
    option: dict[str, Any],
    idx: int,
    source_truth: list[dict[str, Any]] | None = None,
    *,
    is_recommended: bool = False,
) -> dict[str, Any]:
    option_id = f"option_{chr(ord('a') + idx - 1)}"
    dataset = str(option.get("dataset") or "")
    source_label = _source_label(dataset, PROJECT_ROOT)
    sql_label = _sql_column_label(dataset, str(option.get("column") or "unknown"))
    proof = _physical_option_proof(option, source_truth or [])
    score = float(option.get("score") or 0)
    reason_text = str(option.get("reason") or "profile column candidate")
    # Recalibrated to THIS function's own weight scale (see
    # _profile_candidate_score: exact match +100, partial match +60, dataset
    # alignment +30, generic KPI-text containment +20, ID/generic-column
    # penalty -30) -- the previous 6/3 bar was tuned for a different scorer
    # entirely and let a single generic containment hit (+20) pass as "high".
    # A bare generic-containment-only score must cap at medium: name
    # similarity/context overlap alone is never sufficient evidence on its own.
    confidence = "high" if score >= 60 else ("medium" if score >= 20 else "low")
    samples = list(option.get("observed_values") or [])[:3]
    sample_phrase = (
        f"Sample values: {', '.join(repr(s) for s in samples)}."
        if samples else "No sample values recorded for this column."
    )
    summary = (
        f"`{sql_label}` — {reason_text}. {sample_phrase} "
        f"Source: `{source_label}`."
    )
    if is_recommended:
        summary = "RECOMMENDED. " + summary
    return {
        "option_id": option_id,
        "label": sql_label,
        "business_summary": summary,
        "evidence_summary": reason_text,
        "json_backed": True,
        "evidence_state": option.get("evidence_state"),
        "confidence": confidence,
        "confidence_score": round(score, 2),
        "is_recommended": is_recommended,
        "needs_user_confirmation": True,
        "physical_column_option": option,
        "proof_packet": proof,
        "sql_preference": "sql",
        "query": proof["query"],
        "result_demo_table": proof["result_demo_table"],
    }


def _physical_option_proof(option: dict[str, Any], source_truth: list[dict[str, Any]]) -> dict[str, Any]:
    column = str(option.get("column") or "")
    dataset = str(option.get("dataset") or "")
    sql_table = _sql_table_label(dataset)
    samples = list(option.get("observed_values") or option.get("value_profile", {}).get("sample_values") or [])[:8]
    query = option.get("answer_demo", {}).get("query") or (
        f'SELECT "{column}", COUNT(*) AS row_count\n'
        f'FROM "{dataset_display_stem(dataset) or "source_table"}"\n'
        f'GROUP BY "{column}"\n'
        "ORDER BY row_count DESC\n"
        "LIMIT 10;"
    )
    return {
        "proof_type": "physical_column_candidate",
        "kpi_source_truth": source_truth,
        "required_columns": [
            {
                "business_field": str(source_truth[0]["business_question"] if source_truth else option.get("kpi_name") or ""),
                "sql_table": sql_table,
                "sql_column": _sql_column_label(dataset, column),
                "physical_column": column,
                "dataset": dataset,
                "profile_path": option.get("profile_path"),
                "dtype": option.get("dtype"),
                "row_count": option.get("row_count"),
                "evidence_state": option.get("evidence_state"),
                "sample_values": samples,
            }
        ],
        "formula": "",
        "derived_logic": "",
        "proof_sources": option.get("semantic_meaning_sources") or option.get("mapping_proof", {}).get("source_files") or [],
        "sql_preference": "sql",
        "query": query,
        "result_demo_table": option.get("answer_demo", {}).get("sample_output_table") or _markdown_table(
            [{column or "value": value, "row_count": "<computed>"} for value in samples[:5]]
        ),
        "demo_note": option.get("answer_demo", {}).get("note")
        or "Demo rows show expected query output shape from profile samples, not full aggregate results.",
    }


def _answer_demo(
    kpi: dict[str, Any],
    feature: dict[str, Any],
    selected_column: dict[str, Any],
) -> dict[str, Any]:
    dataset = str(selected_column.get("dataset") or "")
    table = dataset_display_stem(dataset) or "source_table"
    column = str(selected_column.get("column") or "")
    feature_label = _slug(str(feature.get("feature") or column))
    metric_text = " ".join(
        str(value or "")
        for value in [kpi.get("name"), kpi.get("metric"), kpi.get("cuts")]
    ).lower()
    same_dataset_features = [
        other
        for other in kpi.get("features", [])
        for source in other.get("source_columns") or []
        if str(source.get("dataset") or "") == dataset and source.get("column")
    ]
    from core.onboarding.lexicon.vocabulary import GENERIC_FINANCIAL_SEED as _FIN_SEED
    cost_column = _first_matching_source_column(kpi, dataset, set(_FIN_SEED))
    select_items = [f'"{column}" AS "{feature_label}"']
    group_by = f'"{column}"'
    order_by = f'"{feature_label}"'
    if "top" in metric_text or "frequent" in metric_text or "number of times" in metric_text:
        select_items.append("COUNT(*) AS row_count")
        order_by = "row_count DESC"
    if cost_column:
        cost_alias = "average_value"
        select_items.append(f'AVG("{cost_column.get("column")}") AS {cost_alias}')
        if order_by == f'"{feature_label}"' and "highest" in metric_text:
            order_by = f"{cost_alias} DESC"
    query = (
        "SELECT\n  "
        + ",\n  ".join(select_items)
        + f'\nFROM "{table}"\nGROUP BY {group_by}\nORDER BY {order_by}\nLIMIT 5;'
    )
    output_rows = _demo_output_rows(selected_column, cost_column, feature_label, metric_text)
    source_samples = [_source_column_sample(selected_column)]
    if cost_column:
        source_samples.append(_source_column_sample(cost_column))
    for other in same_dataset_features:
        if len(source_samples) >= 4:
            break
        source = (other.get("source_columns") or [{}])[0]
        key = (source.get("dataset"), source.get("column"))
        if key not in {(item.get("dataset"), item.get("column")) for item in source_samples}:
            source_samples.append(_source_column_sample(source))
    return {
        "demo_type": "profile_sample_query_demo",
        "kpi_id": kpi.get("kpi_id"),
        "kpi_name": kpi.get("name"),
        "query": query,
        "sample_output": output_rows,
        "sample_output_table": _markdown_table(output_rows),
        "source_column_samples": source_samples,
        "note": (
            "Sample output is built from profile sample values to show the query shape. "
            "It is not the full aggregate result."
        ),
    }


def _first_matching_source_column(
    kpi: dict[str, Any],
    dataset: str,
    terms: set[str],
) -> dict[str, Any] | None:
    for feature in kpi.get("features", []):
        feature_name = _norm(str(feature.get("feature") or ""))
        for source in feature.get("source_columns") or []:
            column_name = _norm(str(source.get("column") or ""))
            if str(source.get("dataset") or "") == dataset and (
                feature_name in terms or any(term in column_name for term in terms)
            ):
                return source
    return None


def _demo_output_rows(
    selected_column: dict[str, Any],
    cost_column: dict[str, Any] | None,
    feature_label: str,
    metric_text: str,
) -> list[dict[str, Any]]:
    values = list(selected_column.get("observed_values") or [])
    if not values:
        values = [row.get(_norm(str(selected_column.get("column") or ""))) for row in (selected_column.get("mapping_proof") or {}).get("sample_output") or []]
    cost_values = list((cost_column or {}).get("observed_values") or [])
    rows = []
    for idx, value in enumerate(values[:5]):
        row: dict[str, Any] = {feature_label: value}
        if "top" in metric_text or "frequent" in metric_text or "number of times" in metric_text:
            row["row_count"] = "<computed from grouped rows>"
        if cost_column:
            row["average_value"] = cost_values[idx] if idx < len(cost_values) else "<computed average>"
        rows.append(row)
    return rows


def _source_column_sample(column: dict[str, Any]) -> dict[str, Any]:
    proof = column.get("mapping_proof") or {}
    sample_output = proof.get("sample_output")
    return {
        "dataset": column.get("dataset"),
        "column": column.get("column"),
        "dictionary_evidence": proof.get("dictionary_evidence"),
        "sample_query": proof.get("sample_query"),
        "sample_output": sample_output,
        "sample_output_table": _markdown_table(sample_output or []),
    }


def _markdown_table(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    columns = list(rows[0].keys())
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows:
        values = [_table_cell(row.get(column)) for column in columns]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def _table_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ")


def _derived_option_payload(
    option: dict[str, Any],
    idx: int,
    source_truth: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    option_id = f"option_{chr(ord('a') + idx - 1)}"
    proof = _derived_option_proof(option, source_truth or [])
    return {
        "option_id": option_id,
        "label": f"Accept candidate formula from `{option.get('source_pattern_id', 'pattern')}`",
        "business_summary": option.get("business_meaning", ""),
        "formula": option.get("formula", ""),
        "json_backed": True,
        "evidence_state": option.get("evidence_state"),
        "confidence": option.get("confidence"),
        "needs_user_confirmation": option.get("needs_user_confirmation"),
        "derived_feature_option": option,
        "proof_packet": proof,
        "sql_preference": "sql",
        "query": proof["query"],
        "result_demo_table": proof["result_demo_table"],
    }


def _derived_option_proof(option: dict[str, Any], source_truth: list[dict[str, Any]]) -> dict[str, Any]:
    input_columns = [column for column in option.get("input_columns") or [] if isinstance(column, dict)]
    required_columns = [
        {
            "business_field": str(column.get("input_name") or ""),
            "physical_column": str(column.get("column") or ""),
            "dataset": str(column.get("dataset") or ""),
            "profile_path": column.get("profile_path"),
            "dtype": column.get("dtype"),
            "row_count": column.get("row_count"),
            "evidence_state": column.get("evidence_state") or "schema_profile_inferred",
            "sample_values": list(column.get("observed_values") or column.get("value_profile", {}).get("sample_values") or [])[:8],
            "reason": column.get("reason") or "",
        }
        for column in input_columns
    ]
    formula = str(option.get("formula") or "")
    query = _derived_sql_query(option)
    result_rows = _derived_demo_rows(option)
    return {
        "proof_type": "derived_formula_candidate",
        "kpi_source_truth": source_truth,
        "required_columns": required_columns,
        "formula": formula,
        "formula_templates": option.get("formula_templates") or {},
        "derived_logic": option.get("business_meaning") or "",
        "derivation_reasoning": option.get("derivation_reasoning") or {},
        "proof_sources": option.get("evidence_sources") or [],
        "synthetic_example": option.get("example") or {},
        "sql_preference": "sql",
        "query": query,
        "result_demo_table": _markdown_table(result_rows),
        "demo_note": "Demo rows show expected output shape from candidate formula evidence, not full aggregate results.",
    }


def _derived_sql_query(option: dict[str, Any]) -> str:
    formula = str(option.get("formula") or "")
    derived_column = str(option.get("derived_column_name") or "derived_value")
    inputs = [column for column in option.get("input_columns") or [] if isinstance(column, dict)]
    select_columns = [str(column.get("column") or column.get("input_name") or "") for column in inputs]
    select_columns = [column for column in select_columns if column]
    table = "joined_source"
    if inputs:
        table = dataset_display_stem(str(inputs[-1].get("dataset") or "")) or table
    select_exprs = [f'"{column}"' for column in select_columns]
    select_exprs.append(f"{formula} AS \"{derived_column}\"" if formula else f'NULL AS "{derived_column}"')
    return "SELECT\n  " + ",\n  ".join(select_exprs) + f'\nFROM "{table}"\nLIMIT 10;'


def _derived_demo_rows(option: dict[str, Any]) -> list[dict[str, Any]]:
    example = option.get("example") if isinstance(option.get("example"), dict) else {}
    input_values = example.get("input") if isinstance(example.get("input"), dict) else {}
    output_values = example.get("output") if isinstance(example.get("output"), dict) else {}
    if input_values or output_values:
        return [{**input_values, **output_values}]
    row: dict[str, Any] = {}
    for column in option.get("input_columns") or []:
        if not isinstance(column, dict):
            continue
        values = column.get("observed_values") or column.get("value_profile", {}).get("sample_values") or []
        row[str(column.get("column") or column.get("input_name") or "")] = values[0] if values else ""
    if row:
        row[str(option.get("derived_column_name") or "derived_value")] = "<computed>"
        return [row]
    return []


def _custom_rule_option(feature: str) -> dict[str, Any]:
    return {
        "option_id": "custom",
        "label": "Enter a custom definition",
        "business_summary": f"Provide a different accepted rule for `{feature}`.",
        "expected_answer_shape": {
            "feature": feature,
            "formula": "",
            "input_columns": [],
            "evidence_source": "",
            "reason": "",
        },
        "json_backed": False,
    }


CLI_AGENT_EVIDENCE_COLUMN_CAP = 60
CLI_AGENT_EVIDENCE_SAMPLE_CAP = 5
CLI_AGENT_DICTIONARY_EXCERPT_CHARS = 2_000
CLI_AGENT_DICTIONARY_DOCUMENT_CAP = 6


def _cli_agent_evidence_pack(
    feature: str,
    items: list[dict[str, Any]],
    workspace: Path,
    repo_root: Path,
) -> dict[str, Any]:
    """Bounded evidence pack the orchestrating CLI agent reads to propose a mapping.

    No raw datasets, no full profile payloads. Only:
      - the feature term and the KPI context that needs it;
      - up to ``CLI_AGENT_EVIDENCE_COLUMN_CAP`` profile columns across all
        datasets, each with dtype, dataset path, and up to
        ``CLI_AGENT_EVIDENCE_SAMPLE_CAP`` sample values;
      - any workspace_feature_definitions already accepted, as worked examples;
      - other features already resolved in kpi_feature_mapping, so the agent can
        see the workspace's resolution style.
    """
    profile_index = _load_json(
        workspace / "interns" / "generated" / "profiles" / "profile_index.json"
    )
    profiles = profile_index.get("profiles") if isinstance(profile_index.get("profiles"), list) else []
    available_columns: list[dict[str, Any]] = []
    layout = WorkspaceLayout(workspace)
    for profile in profiles:
        dataset_raw = str(profile.get("path") or "")
        if not dataset_raw:
            continue
        try:
            if not layout.is_dataset_allowed(Path(dataset_raw)):
                continue
        except OSError:
            continue
        dataset_rel = _rel(Path(dataset_raw), repo_root)
        for column in profile.get("columns") or []:
            if len(available_columns) >= CLI_AGENT_EVIDENCE_COLUMN_CAP:
                break
            name = str(column.get("name") or "")
            if not name:
                continue
            sample_values = _sample_values(column.get("sample_values"))[:CLI_AGENT_EVIDENCE_SAMPLE_CAP]
            available_columns.append(
                {
                    "dataset": dataset_rel,
                    "column": name,
                    "dtype": column.get("dtype"),
                    "null_count": column.get("null_count"),
                    "sample_values": sample_values,
                    "profile_path": profile.get("profile_path"),
                }
            )
        if len(available_columns) >= CLI_AGENT_EVIDENCE_COLUMN_CAP:
            break

    prior_definitions = _load_json(
        workspace / "interns" / "generated" / "contracts" / "workspace_feature_definitions.json"
    )
    prior_accepted: list[dict[str, Any]] = []
    raw_defs = prior_definitions.get("definitions") or prior_definitions.get("workspace_definitions") or {}
    if isinstance(raw_defs, dict):
        for term, definition in raw_defs.items():
            if not isinstance(definition, dict):
                continue
            prior_accepted.append(
                {
                    "feature": term,
                    "source_columns": definition.get("source_columns") or [],
                    "formula": definition.get("formula") or definition.get("expression") or "",
                    "evidence_note": definition.get("evidence_note") or "",
                    "state": definition.get("state") or "user_confirmed",
                }
            )

    feature_mapping = _load_json(
        workspace / "interns" / "generated" / "contracts" / "kpi_feature_mapping.json"
    )
    prior_resolved: list[dict[str, Any]] = []
    for kpi in feature_mapping.get("kpis") or []:
        if not isinstance(kpi, dict):
            continue
        for resolved_feature in kpi.get("features") or []:
            if not isinstance(resolved_feature, dict):
                continue
            if str(resolved_feature.get("state") or "") not in READY_STATES:
                continue
            term = str(resolved_feature.get("feature") or "")
            if not term or term == feature:
                continue
            columns = [
                {
                    "dataset": col.get("dataset"),
                    "column": col.get("column"),
                }
                for col in resolved_feature.get("source_columns") or []
                if isinstance(col, dict)
            ]
            prior_resolved.append({"feature": term, "source_columns": columns})

    kpi_context = []
    seen_kpis: set[str] = set()
    for item in items:
        kpi = item["kpi"]
        kpi_id = str(kpi.get("kpi_id") or "")
        if kpi_id in seen_kpis:
            continue
        seen_kpis.add(kpi_id)
        kpi_context.append(
            {
                "kpi_id": kpi_id,
                "name": kpi.get("name"),
                "description": kpi.get("description"),
                "metric": kpi.get("metric"),
                "cuts": kpi.get("cuts"),
            }
        )

    # Data-dictionary excerpts: extracted by methodology_parser during
    # onboarding from any PDF/DOCX data-model files. Bounded excerpts only —
    # the CLI agent gets meaning context, not full PHI/PII text dumps.
    dictionary_excerpts: list[dict[str, Any]] = []
    dictionary_index = _load_json(
        workspace / "interns" / "generated" / "data_dictionary" / "index.json"
    )
    for doc in (dictionary_index.get("documents") or [])[:CLI_AGENT_DICTIONARY_DOCUMENT_CAP]:
        if not isinstance(doc, dict):
            continue
        text_rel = doc.get("text_path")
        if not isinstance(text_rel, str) or not text_rel:
            continue
        text_path = (repo_root / text_rel).resolve()
        if not text_path.exists():
            continue
        try:
            raw = text_path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        if not raw.strip():
            continue
        excerpt = raw[:CLI_AGENT_DICTIONARY_EXCERPT_CHARS]
        dictionary_excerpts.append(
            {
                "source_document": doc.get("path"),
                "text_path": text_rel,
                "char_count": doc.get("char_count"),
                "excerpt": excerpt,
                "excerpt_truncated": len(raw) > CLI_AGENT_DICTIONARY_EXCERPT_CHARS,
            }
        )

    return {
        "feature": feature,
        "kpi_context": kpi_context,
        "available_columns": available_columns,
        "prior_accepted_definitions": prior_accepted[:20],
        "prior_resolved_features": prior_resolved[:20],
        "data_dictionary_excerpts": dictionary_excerpts,
        "evidence_caps": {
            "columns": CLI_AGENT_EVIDENCE_COLUMN_CAP,
            "samples_per_column": CLI_AGENT_EVIDENCE_SAMPLE_CAP,
            "prior_examples": 20,
            "dictionary_documents": CLI_AGENT_DICTIONARY_DOCUMENT_CAP,
            "dictionary_chars_per_document": CLI_AGENT_DICTIONARY_EXCERPT_CHARS,
        },
        "no_raw_data_policy": (
            "Propose only from columns listed in available_columns. Do not invent column names. "
            "Do not request raw dataset reads. If evidence is insufficient, return option_b (custom). "
            "Data-dictionary excerpts are bounded and may contain partial context; ground every "
            "proposed mapping in a profile column."
        ),
    }


def _cli_agent_task_text(
    feature: str,
    applies_to: list[str],
    evidence_pack: dict[str, Any],
    workspace: Path,
    repo_root: Path,
) -> dict[str, Any]:
    workspace_rel = _rel(workspace, repo_root)
    apply_template = (
        f"uv run apply-kpi-panel-answer --workspace {workspace_rel} "
        f"--domain <domain> --answer custom --custom-definition '<JSON proposal>' "
        f"--via-cli-agent"
    )
    confirm_template = (
        f"uv run confirm-cli-agent-proposal --workspace {workspace_rel} "
        f"--feature {feature} --decision confirm"
    )
    reject_template = (
        f"uv run confirm-cli-agent-proposal --workspace {workspace_rel} "
        f"--feature {feature} --decision reject"
    )
    return {
        "for_cli_agent": True,
        "instruction": (
            f"You are the orchestrating CLI agent. Feature `{feature}` is unresolved for "
            f"{len(applies_to)} KPI(s). Read `cli_agent_evidence_pack` and propose a single JSON "
            "mapping in the shape of `options[0].expected_answer_shape`. Present the proposal to "
            "the user in chat with the columns you chose and the reason. Apply with "
            "`--via-cli-agent` so the decision is recorded as `cli_agent_proposed` (NOT "
            "`user_confirmed`), then ask the user to run `confirm-cli-agent-proposal` to "
            "finalize."
        ),
        "agent_steps": [
            "Read the cli_agent_evidence_pack section in current.json.",
            "Choose a column (or formula over multiple columns) from available_columns ONLY.",
            "Fill option_a.expected_answer_shape with: feature, source_columns, formula or "
            "expression if derived, evidence_source pointing to the profile_path used, and a "
            "one-sentence reason that names the evidence you relied on.",
            "Show the JSON to the user along with the columns and their sample values.",
            "Wait for explicit user approval ('yes', 'apply it', or a modification request).",
            f"On approval, run: {apply_template}",
            f"After the user reviews the resulting mapping, run: {confirm_template}",
            f"If the user rejects after seeing the recorded proposal, run: {reject_template}",
        ],
        "do_not": [
            "Do not invent column names that are not in available_columns.",
            "Do not request access to raw datasets.",
            "Do not call apply-kpi-panel-answer before user approval.",
            "Do not omit `--via-cli-agent` when applying — that skips the user-confirmation step.",
            "Do not run confirm-cli-agent-proposal yourself without explicit user direction.",
            "Do not summarize away the evidence pack when presenting to the user.",
        ],
        "apply_command_template": apply_template,
        "confirm_command_template": confirm_template,
        "reject_command_template": reject_template,
    }


def _cli_agent_proposal_option(
    feature: str,
    evidence_pack: dict[str, Any],
) -> dict[str, Any]:
    return {
        "option_id": "option_a",
        "label": "Agent proposes a mapping from bounded evidence",
        "business_summary": (
            f"The orchestrating CLI agent reads the evidence pack and proposes a mapping for "
            f"`{feature}`. User approval is required before the apply step runs."
        ),
        "expected_answer_shape": {
            "feature": feature,
            "source_columns": [
                {"dataset": "<one of evidence_pack.available_columns[].dataset>",
                 "column": "<one of evidence_pack.available_columns[].column>"}
            ],
            "formula": "<optional expression over the source columns; empty for direct mapping>",
            "evidence_source": "<profile_path used to justify the choice>",
            "reason": "<one sentence naming the evidence relied on>",
            "agent_confidence": "<low|medium|high>",
            "needs_user_confirmation": True,
        },
        "json_backed": True,
        "evidence_column_count": len(evidence_pack.get("available_columns") or []),
        "evidence_caps": evidence_pack.get("evidence_caps"),
    }


def _evidence_files(items: list[dict[str, Any]], repo_root: Path) -> list[dict[str, str]]:
    # Keyed by normalised path so each file appears once with its first-seen purpose.
    seen: dict[str, str] = {}

    def _add(raw: str, purpose: str) -> None:
        if not raw:
            return
        p = Path(raw)
        key = str(_rel(p, repo_root) if p.is_absolute() else raw)
        if key not in seen:
            seen[key] = purpose

    for item in items:
        source = item["kpi"].get("source")
        if source:
            _add(str(source), "kpi_source")
        for option in item["feature"].get("derived_feature_options") or []:
            for source_item in option.get("evidence_sources") or []:
                file = source_item.get("file")
                if file:
                    _add(str(file), "derivation_evidence")
            for column in option.get("input_columns") or []:
                profile = column.get("profile_path")
                if profile:
                    _add(str(profile), "input_column_profile")
        for column in item["feature"].get("source_columns") or []:
            profile = column.get("profile_path")
            if profile:
                _add(str(profile), "source_column_profile")
            proof = column.get("mapping_proof") or {}
            for file in proof.get("source_files") or []:
                if file:
                    _add(str(file), "mapping_proof")
            for meaning in column.get("semantic_meaning_sources") or []:
                file = meaning.get("file")
                if file:
                    _add(str(file), "semantic_meaning_source")
    return sorted(
        [{"file": k, "purpose": v} for k, v in seen.items()],
        key=lambda x: (x["purpose"], x["file"]),
    )


def _blocked_kpi_details(mapping: dict[str, Any]) -> list[dict[str, Any]]:
    """Per-KPI definition asks for the blocked-without-feature-question card.

    Each blocked KPI contributes its name, prose excerpt, open questions, and
    the workspace-evidence anchors its definition blocker carries, so the
    definition-help card asks concretely per KPI instead of a bare command.
    """
    details: list[dict[str, Any]] = []
    for kpi in mapping.get("kpis", []) or []:
        if not isinstance(kpi, dict) or kpi.get("status") != "blocked_questions_pending":
            continue
        prose = ""
        anchors: list[dict[str, Any]] = []
        for feature in kpi.get("features") or []:
            if not isinstance(feature, dict):
                continue
            for evidence in feature.get("evidence") or []:
                if not isinstance(evidence, dict):
                    continue
                if evidence.get("type") == "kpi_prose" and not prose:
                    prose = str(evidence.get("excerpt") or "")
                elif evidence.get("type") == "prose_term_match":
                    anchors.append(
                        {
                            "dataset": evidence.get("source", ""),
                            "column": evidence.get("column", ""),
                            "matched_terms": evidence.get("matched_terms") or [],
                            "dictionary_description": evidence.get(
                                "dictionary_description", ""
                            ),
                        }
                    )
        details.append(
            {
                "kpi_id": kpi.get("kpi_id", ""),
                "name": kpi.get("name", ""),
                "prose_excerpt": prose,
                "open_questions": list(kpi.get("open_questions") or []),
                "anchor_evidence": anchors[:6],
            }
        )
    return details


def _empty_panel(mapping: dict[str, Any], workspace: Path, repo_root: Path) -> dict[str, Any]:
    summary = mapping.get("summary", {}) or {}
    blocked_count = 0
    try:
        blocked_count = int(summary.get("blocked_kpi_count") or 0)
    except (TypeError, ValueError):
        blocked_count = 0

    panel = {
        "artifact_type": "blocker_question_panel/current.json",
        "version": PANEL_VERSION,
        "generated_by": "blocker-question-panel",
        "workspace": _rel(workspace, repo_root),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "question_id": "none",
        "feature": "",
        "status": "no_blocker_questions",
        "blocker": "No unresolved blocker questions were found.",
        "question": "",
        "answer_type": "none",
        "recommended_answer": "",
        "why": "",
        "options": [],
        "interaction_contract": INTERACTION_CONTRACT,
        "summary": summary,
    }

    # Dead-end guard: KPIs are still blocked, yet no answerable question was
    # produced. Don't go silent -- carry the stage routing (specialists + skills)
    # so the orchestrator activates the right help instead of looping. Generic:
    # the routing roster comes from delegation.STAGE_ROUTING, not a domain list.
    if blocked_count > 0:
        panel["status"] = "blocked_without_question"
        panel["blocker"] = (
            f"{blocked_count} KPI(s) are blocked but produced no answerable feature "
            "question -- likely an incomplete/undefined definition rather than a "
            "feature-mapping gap. This needs definition help, not a column choice."
        )
        panel["question"] = (
            "Define the metric and grain for the blocked KPI(s) with "
            "`uv run apply-kpi-definition --workspace <workspace> --kpi-id "
            "<kpi_id> --metric \"...\" --cuts \"...\" --confirmed-by \"<name>\"` "
            "(one per KPI), then re-run prepare-kpi-blocker-panel. "
            "Or restart KPI generation."
        )
        panel["blocked_kpi_details"] = _blocked_kpi_details(mapping)
        try:
            from core.onboarding.workspace.delegation import routing_for

            roster = routing_for("kpi_definition")
            if roster.get("agents"):
                panel["required_specialists"] = roster["agents"]
            if roster.get("skills"):
                panel["suggested_skills"] = [
                    {"name": s, "why": "blocked KPI with no answerable question"}
                    for s in roster["skills"]
                ]
        except Exception:  # pragma: no cover - routing is advisory
            pass
    return panel


def _render_all_blockers_overview(questions: list[dict[str, Any]], workspace_rel: str) -> str:
    """The full open-decision surface: a summary table of every blocker plus each
    one's compact card. Lets an operator see and plan ALL pending decisions, not
    just the single `current` one. Generic across answer types."""
    ws_label = workspace_rel.split("/")[-1] if workspace_rel else ""
    n = len(questions)
    lines = [
        f"# Blocker Panel — {n} open decision{'s' if n != 1 else ''} ({ws_label})",
        "",
    ]
    if not questions:
        lines.append("No open blockers. All features are resolved.")
        return "\n".join(lines) + "\n"
    lines += [
        "Answer any or all below. Some auto-resolve once a related one is answered",
        "(workspace-level mappings + KPI definitions cascade), so the panel",
        "re-derives after each answer. [REC] marks the recommended default.",
        "",
        "| # | Decision | Affects | Type | Recommended |",
        "|---|----------|---------|------|-------------|",
    ]
    for i, q in enumerate(questions, start=1):
        rec = str(q.get("recommended_option_id") or "")
        opts = {str(o.get("option_id")): o for o in q.get("options") or []}
        rec_label = str((opts.get(rec) or {}).get("label") or rec) or "—"
        affects = ", ".join(str(k) for k in q.get("applies_to_kpis") or []) or "—"
        lines.append(
            f"| {i} | `{q.get('feature', '')}` | {affects} | "
            f"{q.get('answer_type', '')} | {rec}: {rec_label} |"
        )
    lines.append("")
    for i, q in enumerate(questions, start=1):
        lines.append(f"## {i}. {q.get('feature', '')}")
        lines.append("")
        lines.append(_render_markdown_compact(q))
    return "\n".join(lines) + "\n"


def _render_markdown_compact(panel: dict[str, Any]) -> str:
    """Compact, decision-first blocker card written to ``current.md``.

    Leads with the affected KPI(s), the decision, and the options -- the only
    things a human needs to answer the blocker. The full evidence/proof render
    (feature-resolution tables, source-truth packets, per-option proof packets,
    raw JSON) lives in ``current_full.md``; the machine contract lives in
    ``current.json``. Stays generic across every ``answer_type`` because it only
    reads fields that every branch of ``_question_for_cluster`` populates.
    """
    feature = panel.get("feature") or "None"
    workspace = str(panel.get("workspace") or "")
    ws_label = workspace.split("/")[-1] if workspace else ""
    applies = panel.get("applies_to_kpis") or []
    recommended_id = str(panel.get("recommended_option_id") or "")
    lines = [f"# Blocker: {feature}", ""]

    context_bits = []
    if ws_label:
        context_bits.append(f"Workspace: {ws_label}")
    if applies:
        context_bits.append("Affects: " + ", ".join(str(k) for k in applies))
    if panel.get("answer_type"):
        context_bits.append(f"Type: {panel.get('answer_type')}")
    if context_bits:
        lines.extend([" · ".join(context_bits), ""])

    # Affected KPI truth -- the "exactly what's needed" anchor, one tight block.
    source_truth = panel.get("kpi_source_truth") or []
    for truth in source_truth:
        kpi_id = str(truth.get("kpi_id") or "")
        # business_question is raw prose typed by a business user into the
        # workspace's own source workbook -- untrusted, and this is the
        # panel's declared primary_artifact (current.md), read verbatim by
        # the orchestrating CLI agent.
        question = neutralize_text(str(truth.get("business_question") or ""))
        lines.append(f"**{kpi_id}** — {question}" if kpi_id else question)
        meta = []
        metric = str(truth.get("metric") or "")
        cuts = ", ".join(str(c) for c in (truth.get("cuts") or []))
        if metric:
            meta.append(f"Metric: `{metric}`")
        if cuts:
            meta.append(f"Cuts: {cuts}")
        if meta:
            lines.append("  " + "   ".join(meta))
    if source_truth:
        lines.append("")

    # The decision: blocker (why) + question (what to answer).
    lines.extend(["## Decide", ""])
    if panel.get("blocker"):
        lines.extend([str(panel.get("blocker")), ""])
    lines.extend([str(panel.get("question") or ""), ""])

    # Options: one identity line + at most one reason line + one samples line.
    lines.extend(["## Options", ""])
    for idx, option in enumerate(panel.get("options") or [], start=1):
        option_id = str(option.get("option_id") or "")
        label = str(option.get("label") or "")
        is_recommended = bool(
            option.get("is_recommended")
            or (recommended_id and option_id == recommended_id)
        )
        marker = " [RECOMMENDED]" if is_recommended else ""
        confidence = str(option.get("confidence") or "")
        conf_text = f"  (confidence: {confidence})" if confidence else ""
        lines.append(f"{idx}.{marker} `{option_id}` · {label}{conf_text}")
        reason = _compact_option_reason(option)
        if reason:
            lines.append(f"   {reason}")
        samples = _compact_option_samples(option)
        if samples:
            # observed_values/sample_values are raw column content from the
            # workspace's own data -- untrusted.
            lines.append("   Samples: " + ", ".join(neutralize_text(str(s)) for s in samples))
    lines.append("")

    # How to answer.
    apply_target = recommended_id or "<option_id>"
    apply_ws = workspace or "<workspace>"
    lines.extend(
        [
            "## Answer",
            "",
            "```text",
            f"uv run apply-kpi-panel-answer --workspace {apply_ws} --domain <domain> --answer {apply_target}",
            "```",
            "",
            "Full proof & evidence: `current_full.md` · machine contract: `current.json`.",
        ]
    )
    return "\n".join(lines) + "\n"


def _compact_option_reason(option: dict[str, Any]) -> str:
    """One short reason line for an option, without the proof-packet dump."""
    reason = str(option.get("evidence_summary") or "").strip()
    if not reason:
        summary = str(option.get("business_summary") or "").strip()
        if summary.startswith("RECOMMENDED. "):
            summary = summary[len("RECOMMENDED. "):]
        # Drop sample/source boilerplate the compact view re-derives below.
        for token in (" Sample values:", " Samples:", " Source:"):
            pos = summary.find(token)
            if pos != -1:
                summary = summary[:pos]
        reason = summary.split(". ")[0].strip().rstrip(".")
    return reason


def _compact_option_samples(option: dict[str, Any], limit: int = 3) -> list[Any]:
    """Up to ``limit`` sample values for an option, from whichever evidence
    shape it carries (physical column, proof packet, or derived feature)."""
    physical = option.get("physical_column_option") or {}
    samples = list(physical.get("observed_values") or [])
    if not samples:
        proof = option.get("proof_packet") or {}
        required = proof.get("required_columns") or []
        if required:
            samples = list((required[0] or {}).get("sample_values") or [])
    if not samples:
        derived = option.get("derived_feature_option") or {}
        samples = list(derived.get("observed_values") or [])
    return samples[:limit]


def _render_markdown(panel: dict[str, Any]) -> str:
    allowed_option_ids = ", ".join(
        f"`{option.get('option_id', '')}`" for option in panel.get("options") or []
    )
    lines = [
        f"# Blocker Question Panel: {panel.get('feature') or 'None'}",
        "",
        f"- Workspace: `{panel.get('workspace', '')}`",
        f"- Applies to KPIs: {', '.join(panel.get('applies_to_kpis') or []) or 'none'}",
        f"- Reuse scope: `{panel.get('reuse_scope', 'none')}`",
        f"- Answer type: `{panel.get('answer_type', '')}`",
        "",
    ]
    if panel.get("feature_resolution_table"):
        lines.extend(_render_feature_resolution_table(panel["feature_resolution_table"]))
    if panel.get("sample_evidence"):
        lines.extend(_render_sample_evidence(panel["sample_evidence"]))
    if panel.get("kpi_preview"):
        lines.extend(_render_kpi_preview(panel["kpi_preview"]))
    interaction = panel.get("interaction_contract") or {}
    if interaction:
        lines.extend(
            [
                "## Interaction Contract",
                "",
                str(interaction.get("instruction", "")),
                "",
                f"- Display mode: `{interaction.get('display_mode', '')}`",
                f"- Generic answer picker allowed: `{interaction.get('generic_answer_picker_allowed')}`",
                "",
            ]
        )
    prior = panel.get("prior_decision_wiki") or {}
    if prior:
        lines.extend(
            [
                "## Prior Decision (from wiki)",
                "",
                f"- Note: `{prior.get('path', '')}`",
                f"- Summary: {prior.get('summary', '')}",
                f"- Last updated: {prior.get('updated', '')}",
                "",
            ]
        )
        if prior.get("has_user_why") and prior.get("user_why"):
            # Human-typed freeform note, read straight off disk -- untrusted.
            lines.extend(
                ["### User-recorded *why*", "", neutralize_text(str(prior.get("user_why"))), ""]
            )
    if panel.get("kpi_source_truth"):
        lines.extend(["## KPI Source Truth", ""])
        for truth in panel.get("kpi_source_truth") or []:
            lines.extend(
                [
                    f"### {truth.get('kpi_id', '')}",
                    "",
                    # Raw workbook prose typed by a business user -- untrusted.
                    f"- Business question: {neutralize_text(str(truth.get('business_question', '')))}",
                    f"- Description: {neutralize_text(str(truth.get('description', '')))}",
                    f"- Metric from source: `{truth.get('metric', '')}`",
                    f"- Cuts / dimensions from source: {', '.join(truth.get('cuts') or [])}",
                    f"- Source: `{truth.get('source', '')}`",
                ]
            )
            if truth.get("source_sheet") or truth.get("source_cells"):
                lines.append(f"- Sheet: `{truth.get('source_sheet', '')}`")
                cells = truth.get("source_cells") or {}
                cell_text = ", ".join(f"{key}={value}" for key, value in cells.items() if value)
                if cell_text:
                    lines.append(f"- Source cells: {cell_text}")
            lines.extend(["", str(truth.get("source_truth_note", "")), ""])
    if panel.get("panel_contract"):
        contract = panel.get("panel_contract") or {}
        lines.extend(
            [
                "## Panel Contract",
                "",
                f"- Display shape: `{contract.get('display_shape', '')}`",
                f"- Default code preference: `{contract.get('default_code_preference', panel.get('default_code_preference', 'sql'))}`",
                "- Required sections: "
                + ", ".join(f"`{item}`" for item in contract.get("must_include", []) or []),
                "",
            ]
        )
    if panel.get("output_dialect"):
        dialect = panel.get("output_dialect") or {}
        lines.extend(
            [
                "## Output Dialect",
                "",
                f"- Default: `{dialect.get('label', 'SQL (default)')}`",
                "- Alternatives: " + ", ".join(f"`{item}`" for item in dialect.get("alternatives", [])),
                f"- Rule: {dialect.get('rule', '')}",
                "",
            ]
        )
    if panel.get("immutable_kpi_policy"):
        policy = panel.get("immutable_kpi_policy") or {}
        lines.extend(
            [
                "## Immutable KPI Policy",
                "",
                str(policy.get("rule", "")),
                "",
                f"- Understanding is review context only: `{policy.get('understanding_is_review_context_only')}`",
                f"- Placeholder SQL is non-executable: `{policy.get('placeholder_sql_is_non_executable')}`",
                "",
            ]
        )
    if panel.get("kpi_understanding"):
        lines.extend(["## KPI Understanding Review", ""])
        for item in panel.get("kpi_understanding") or []:
            original = item.get("original_kpi") or {}
            lines.extend(
                [
                    f"### {item.get('kpi_id', '')}",
                    "",
                    f"- Presentation level: `{item.get('presentation_level', '')}`",
                    f"- Requires understanding approval: `{item.get('requires_understanding_approval')}`",
                    f"- Affected unresolved feature: `{item.get('affected_unresolved_feature', '')}`",
                    f"- Original KPI: {neutralize_text(str(original.get('business_question', '')))}",
                    f"- Source metric: `{original.get('metric', '')}`",
                    f"- Source cuts / filters: {', '.join(original.get('cuts') or [])}",
                    "",
                    "#### My Understanding",
                    "",
                    str(item.get("my_understanding", "")),
                    "",
                    str(item.get("understanding_warning", "")),
                    "",
                    f"#### Output Dialect: {item.get('output_dialect', 'SQL')}",
                    "",
                    "#### Strict Proven SQL",
                    "",
                    "```sql",
                    str(item.get("strict_proven_sql", "")),
                    "```",
                    "",
                    "#### Placeholder Intent SQL",
                    "",
                    "```sql",
                    str(item.get("intent_sql_sketch", "")),
                    "```",
                    "",
                    "#### Demo Result Table",
                    "",
                    str(item.get("demo_result_table", "")),
                    "",
                ]
            )
    lines += [
        "## Required User-Facing Ask",
        "",
        "Use this section when asking the user for the blocker answer. Do not replace it with a freehand summary.",
        "",
        f"- Question: {panel.get('question', '')}",
        f"- Recommended option id: `{panel.get('recommended_option_id', '')}`",
        f"- Recommended answer: {panel.get('recommended_answer', '')}",
        f"- Allowed option ids: {allowed_option_ids}",
        "",
        "Do not state that another option is recommended unless `current.json` says so.",
        "",
        "## Blocker",
        "",
        str(panel.get("blocker", "")),
        "",
        "## Question",
        "",
        str(panel.get("question", "")),
        "",
        "## Instruction",
        "",
        str(panel.get("recommended_answer", "")),
        "",
        "## Why",
        "",
        str(panel.get("why", "")),
        "",
    ]
    blocked_details = panel.get("blocked_kpi_details") or []
    if blocked_details:
        lines += [
            "## Blocked KPIs (definition needed)",
            "",
        ]
        for detail in blocked_details:
            lines.append(f"### {detail.get('kpi_id', '')} -- {detail.get('name', '')}")
            lines.append("")
            # Raw KPI-prose excerpt sourced from the workspace's own
            # documents -- untrusted; whitespace normalization alone isn't
            # sanitization.
            prose = neutralize_text(str(detail.get("prose_excerpt") or "").strip())
            if prose:
                lines.append("> " + " ".join(prose.split())[:400])
                lines.append("")
            for question in detail.get("open_questions") or []:
                lines.append(f"- Ask: {question}")
            anchors = detail.get("anchor_evidence") or []
            if anchors:
                lines.append("- Workspace evidence anchors:")
                for anchor in anchors:
                    dataset = str(anchor.get("dataset") or "")
                    dataset_label = dataset.split("/")[-1] if dataset else ""
                    terms = ", ".join(str(t) for t in anchor.get("matched_terms") or [])
                    gloss = str(anchor.get("dictionary_description") or "")
                    suffix = f" -- {gloss}" if gloss else ""
                    lines.append(
                        f"  - `{dataset_label}.{anchor.get('column', '')}` "
                        f"(matched: {terms}){suffix}"
                    )
            lines.append("")
    cli_agent_task = panel.get("cli_agent_task") or {}
    if cli_agent_task:
        lines.extend(
            [
                "## CLI Agent Task",
                "",
                str(cli_agent_task.get("instruction", "")),
                "",
                "### Agent Steps",
                "",
            ]
        )
        for step in cli_agent_task.get("agent_steps") or []:
            lines.append(f"- {step}")
        lines.append("")
        do_not = cli_agent_task.get("do_not") or []
        if do_not:
            lines.extend(["### Do Not", ""])
            for rule in do_not:
                lines.append(f"- {rule}")
            lines.append("")
        apply_template = cli_agent_task.get("apply_command_template")
        if apply_template:
            lines.extend(
                [
                    "### Apply Command Template (run only after user approval)",
                    "",
                    "```text",
                    str(apply_template),
                    "```",
                    "",
                ]
            )
    evidence_pack = panel.get("cli_agent_evidence_pack") or {}
    if evidence_pack:
        # The bounded evidence pack the orchestrating CLI agent reads to
        # propose a mapping -- contains raw sample_values, human/agent-typed
        # evidence_note freeform text, and (via data_dictionary_excerpts) raw
        # PDF/DOCX text extracted from the workspace's own documents. All of
        # it untrusted. Neutralize every string leaf before serializing,
        # rather than the dumped JSON text (which risks a match spanning
        # JSON syntax and corrupting structure).
        safe_evidence_pack = neutralize_json(evidence_pack)
        lines.extend(
            [
                "## CLI Agent Evidence Pack",
                "",
                f"- Feature: `{evidence_pack.get('feature', '')}`",
                f"- Available columns: {len(evidence_pack.get('available_columns') or [])}",
                f"- Prior accepted definitions: {len(evidence_pack.get('prior_accepted_definitions') or [])}",
                f"- Prior resolved features: {len(evidence_pack.get('prior_resolved_features') or [])}",
                "",
                str(evidence_pack.get("no_raw_data_policy", "")),
                "",
                "```json",
                json.dumps(safe_evidence_pack, indent=2, default=str),
                "```",
                "",
            ]
        )
    lines += [
        "## Options",
        "",
    ]
    for option in panel.get("options") or []:
        lines.extend(
            [
                f"### {option.get('option_id', '')}: {option.get('label', '')}",
                "",
                str(option.get("business_summary", "")),
                "",
            ]
        )
        if option.get("formula"):
            lines.extend(["```sql", str(option["formula"]), "```", ""])
        proof = option.get("proof_packet") or {}
        if proof:
            lines.extend(_render_option_proof(proof))
        if option.get("executed_sample"):
            lines.extend(_render_executed_sample(option["executed_sample"]))
        if option.get("json_backed"):
            evidence = option.get("derived_feature_option") or option.get("physical_column_option") or {}
            lines.extend(["```json", json.dumps(evidence, indent=2, default=str), "```", ""])
    if panel.get("evidence_files"):
        lines.extend(["## Evidence Files", ""])
        for entry in panel["evidence_files"]:
            if isinstance(entry, dict):
                lines.append(f"- `{entry['file']}` ({entry.get('purpose', '')})")
            else:
                lines.append(f"- `{entry}`")
        lines.append("")
    return "\n".join(lines)


def _render_feature_resolution_table(rows: list[dict]) -> list[str]:
    lines = ["## Feature Resolution", ""]
    if not rows:
        lines.extend(["(no features)", ""])
        return lines
    header_columns = ["Feature", "Resolves as", "Where it lands"]
    lines.append("| " + " | ".join(header_columns) + " |")
    lines.append("| " + " | ".join("---" for _ in header_columns) + " |")
    for row in rows:
        values = [
            _table_cell(row.get("feature")),
            _table_cell(row.get("resolves_as")),
            _table_cell(row.get("where_it_lands")),
        ]
        lines.append("| " + " | ".join(values) + " |")
    lines.append("")
    return lines


def _render_sample_evidence(rows: list[dict]) -> list[str]:
    lines = ["## Sample Evidence", ""]
    if not rows:
        lines.extend(["(no samples)", ""])
        return lines
    header_columns = ["Feature", "Column", "First 5 samples"]
    lines.append("| " + " | ".join(header_columns) + " |")
    lines.append("| " + " | ".join("---" for _ in header_columns) + " |")
    for row in rows:
        samples = row.get("first_samples") or []
        # Raw column content -- already PII-redacted upstream (redact_sample_values)
        # but that's a different concern from injection; neutralize separately.
        sample_text = (
            ", ".join(neutralize_text(str(value)) for value in samples) if samples else "(no samples)"
        )
        values = [
            _table_cell(row.get("feature")),
            _table_cell(row.get("column")),
            _table_cell(sample_text),
        ]
        lines.append("| " + " | ".join(values) + " |")
    lines.append("")
    return lines


def _render_kpi_preview(preview: dict) -> list[str]:
    kpi_id = str(preview.get("kpi_id") or "")
    assumed_option_id = str(preview.get("assumed_option_id") or "")
    caption = str(preview.get("caption") or "")
    result = preview.get("preview_result") or {}
    status = str(result.get("status") or "")
    sql = str(result.get("sql") or "")
    duration_ms = result.get("duration_ms")
    error = result.get("error")

    lines = ["## KPI Preview", ""]
    if status == "ok":
        lines.append(
            f"`{kpi_id}` — preview assuming option `{assumed_option_id}` (the recommendation)."
        )
    else:
        lines.append(
            f"`{kpi_id}` — preview assuming option `{assumed_option_id}`."
        )
    lines.append("")
    if status == "ok" and caption:
        lines.extend([caption, ""])
    lines.extend(["```sql", sql, "```", ""])
    if status == "ok":
        rows = result.get("rows") or []
        columns = result.get("columns") or (list(rows[0].keys()) if rows else [])
        if columns:
            lines.append("| " + " | ".join(str(col) for col in columns) + " |")
            lines.append("| " + " | ".join("---" for _ in columns) + " |")
            for row in rows:
                values = [_table_cell(row.get(col)) for col in columns]
                lines.append("| " + " | ".join(values) + " |")
            lines.append("")
        lines.extend([f"> Executed in {duration_ms}ms via DuckDB.", ""])
    elif status == "empty":
        lines.extend(
            [
                f"> Query executed in {duration_ms}ms but returned 0 rows. "
                "Check the filter scope (e.g., quoted literal filters in the KPI cuts) "
                "and the dataset paths.",
                "",
            ]
        )
    else:
        reason = str(error) if error else status or "unknown error"
        lines.extend(
            [
                f"> Preview unavailable: {reason}. The SQL above is still valid — "
                "run it yourself or rerun the panel after fixing the underlying issue.",
                "",
            ]
        )
    return lines


def _render_executed_sample(preview: dict) -> list[str]:
    lines = ["#### Sample (DuckDB, LIMIT 5)", ""]
    status = str(preview.get("status") or "")
    if status == "ok":
        rows = preview.get("rows") or []
        columns = preview.get("columns") or (list(rows[0].keys()) if rows else [])
        if columns:
            lines.append("| " + " | ".join(str(col) for col in columns) + " |")
            lines.append("| " + " | ".join("---" for _ in columns) + " |")
            for row in rows:
                values = [_table_cell(row.get(col)) for col in columns]
                lines.append("| " + " | ".join(values) + " |")
            lines.append("")
        duration_ms = preview.get("duration_ms")
        lines.extend([f"> Executed in {duration_ms}ms.", ""])
    else:
        error = preview.get("error")
        reason = str(error) if error else status or "unknown error"
        lines.extend([f"> Preview unavailable: {reason}.", ""])
    return lines


def _render_option_proof(proof: dict[str, Any]) -> list[str]:
    lines = ["#### Option Proof Packet", ""]
    if proof.get("derived_logic"):
        lines.extend(["Derived / business logic:", "", str(proof.get("derived_logic")), ""])
    if proof.get("formula"):
        lines.extend(["Formula:", "", "```sql", str(proof.get("formula")), "```", ""])
    required = proof.get("required_columns") or []
    if required:
        lines.extend(
            [
                "SQL mapping and source evidence:",
                "",
                _markdown_table(
                    [
                        {
                            "Business field": row.get("business_field", ""),
                            "SQL table": row.get("sql_table") or _sql_table_label(str(row.get("dataset") or "")),
                            "SQL column": row.get("sql_column")
                            or _sql_column_label(
                                str(row.get("dataset") or ""),
                                str(row.get("physical_column") or ""),
                            ),
                            "Source evidence": _source_label(str(row.get("dataset") or ""), PROJECT_ROOT),
                            "Evidence": row.get("evidence_state", ""),
                            "Sample values": ", ".join(str(value) for value in (row.get("sample_values") or [])[:5]),
                        }
                        for row in required
                    ]
                ),
                "",
            ]
        )
    if proof.get("synthetic_example"):
        lines.extend(["Example:", "", "```json", json.dumps(proof.get("synthetic_example"), indent=2, default=str), "```", ""])
    if proof.get("query"):
        lines.extend(
            [
                f"Code preference: `{proof.get('sql_preference', 'sql')}`",
                "",
                "Query for this option:",
                "",
                "```sql",
                str(proof.get("query")),
                "```",
                "",
            ]
        )
    if proof.get("result_demo_table"):
        lines.extend(["Result demo shape:", "", str(proof.get("result_demo_table")), ""])
    if proof.get("demo_note"):
        lines.extend([str(proof.get("demo_note")), ""])
    return lines


def _norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_") or "item"


def _rel(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()



@anchored("blocker-question-panel")
def main(argv: list[str] | None = None) -> int | None:
    parser = argparse.ArgumentParser(
        description="Generate a stakeholder-friendly blocker question panel."
    )
    parser.add_argument("--workspace", required=True, help="Workspace path, for example workspaces/demo")
    parser.add_argument("--repo-root", default=str(PROJECT_ROOT), help="Repository root. Defaults to current directory.")
    parser.add_argument("--mapping", help="Optional path to kpi_feature_mapping.json.")
    parser.add_argument("--out", help="Optional output directory.")
    parser.add_argument("--json", action="store_true", help="Print JSON summary.")
    args = parser.parse_args(argv)

    from core.onboarding.cli_deprecation import (
        announce_deprecated_cli_redirect,
        is_internal_cli_call,
        warn_soft_deprecated_cli,
    )

    stage_only = bool(args.mapping or args.out)
    if stage_only or is_internal_cli_call():
        warn_soft_deprecated_cli(
            "blocker-question-panel",
            prefer="prepare-kpi-blocker-panel",
            reason="the wrapper runs feature resolution, panel build, and validation atomically",
        )
    else:
        announce_deprecated_cli_redirect(
            "blocker-question-panel",
            prefer="prepare-kpi-blocker-panel",
            reason="the wrapper runs feature resolution, panel build, and validation atomically",
        )
        from core.onboarding.kpi.blocker_cli import prepare_main

        return prepare_main(
            ["--workspace", args.workspace, "--repo-root", args.repo_root]
        )

    result = BlockerQuestionPanelBuilder(
        args.repo_root,
        args.workspace,
        mapping_path=args.mapping,
        output_dir=args.out,
    ).run()
    if args.json:
        print(json.dumps(result.summary(), indent=2))
        return None
    print(f"Wrote {result.question_count} blocker question panel(s) to {result.output_dir}")
    print(f"- {result.current_json}")
    print(f"- {result.current_markdown}")
    print(f"- {result.index_json}")
    return None


if __name__ == "__main__":
    raise SystemExit(main())


# ---------------------------------------------------------------------------
# Preview composition helpers (wired into _question_for_cluster at runtime).
#
# These build the new panel sections introduced for the "real-ops-dashboard"
# rendering: a feature_resolution_table summarizing every feature in the KPI,
# a sample_evidence mini-table with the first observed values per feature
# (PII columns redacted in the display), and per-option executed_sample blocks
# carrying the DuckDB result of running each option's query against the
# workspace data.
# ---------------------------------------------------------------------------


_RESOLVED_STATE_LABELS = {
    "proven_direct": "proven_direct",
    "proven_alias": "proven_alias",
    "proven_join": "proven_join",
    "proven_formula": "proven_formula",
    "proven_taxonomy": "proven_taxonomy",
    "user_confirmed": "user_confirmed",
    "cli_agent_proposed": "cli_agent_proposed",
    "blocked_missing_evidence": "blocked_missing_evidence",
    "blocked_ambiguous": "blocked_ambiguous",
    "rejected": "rejected",
}


def _where_it_lands(feature: dict[str, Any]) -> str:
    """One-line description of where a feature maps to in the schema."""

    state = str(feature.get("state") or "")
    sources = feature.get("source_columns") or []
    if state == "proven_join" and sources:
        first = sources[0]
        dataset_stem = dataset_display_stem(str(first.get("dataset") or ""))
        column = str(first.get("column") or "?")
        # Detect a join key by scanning source_columns for an *_id-shaped column
        # that also appears on the parent table. Heuristic but useful.
        for column_ref in sources:
            col_name = str(column_ref.get("column") or "")
            if col_name and col_name.lower().endswith("id"):
                return f"{dataset_stem}.{column} via {col_name}"
        return f"{dataset_stem}.{column} (join)"
    if state.startswith("proven_") or state == "user_confirmed":
        if sources:
            first = sources[0]
            return f"{dataset_display_stem(str(first.get('dataset') or ''))}.{first.get('column') or '?'}"
        formula = feature.get("derived_formula") or feature.get("resolution_type") or ""
        return f"derived ({formula})" if formula else "(resolved)"
    if state == "cli_agent_proposed":
        if sources:
            first = sources[0]
            return f"proposed: {dataset_display_stem(str(first.get('dataset') or ''))}.{first.get('column') or '?'} (awaiting confirmation)"
        return "proposed (awaiting confirmation)"
    # Blocked states: try to surface the most promising direction.
    derived = feature.get("derived_feature_options") or []
    if derived:
        first = derived[0]
        inputs = first.get("input_columns") or []
        if inputs and isinstance(inputs[0], dict):
            return f"derived from {dataset_display_stem(str(inputs[0].get('dataset') or ''))}.{inputs[0].get('column') or '?'}"
        return "derived (no proven inputs)"
    if sources:
        first = sources[0]
        return f"candidate: {dataset_display_stem(str(first.get('dataset') or ''))}.{first.get('column') or '?'}"
    return "(no candidate)"


def _build_feature_resolution_table(mapping: dict[str, Any]) -> list[dict[str, str]]:
    """Build the 3-column feature-resolution table.

    Features are deduplicated by name; the first occurrence's state wins.
    Order: resolved features first (by KPI iteration order), then blocked ones.
    """

    seen: set[str] = set()
    rows: list[dict[str, str]] = []
    for kpi in mapping.get("kpis", []):
        for feature in kpi.get("features", []):
            name = str(feature.get("feature") or "")
            if not name or name in seen:
                continue
            seen.add(name)
            state = str(feature.get("state") or "")
            rows.append(
                {
                    "feature": name,
                    "resolves_as": _RESOLVED_STATE_LABELS.get(state, state or "unknown"),
                    "where_it_lands": _where_it_lands(feature),
                }
            )
    rows.sort(key=lambda r: (0 if r["resolves_as"].startswith("proven_") or r["resolves_as"] == "user_confirmed" else 1, r["feature"]))
    return rows


def _workspace_redaction_patterns(workspace_path: Path | None) -> tuple[str, ...]:
    """Default PII patterns extended with the workspace's user data policy.

    Thin wrapper over the shared ``pii_redaction.workspace_redaction_patterns``
    (single source of truth) so the panel, the result packet, and the verifier
    all redact with the same effective pattern set.
    """
    from core.onboarding.kpi.pii_redaction import workspace_redaction_patterns

    return workspace_redaction_patterns(workspace_path)


def _build_sample_evidence(
    mapping: dict[str, Any], workspace_path: Path | None = None
) -> list[dict[str, Any]]:
    """Build the sample-evidence mini-table.

    For each feature that has source_columns with samples, emit one row with
    `feature`, `column` (dataset_stem.column), and the first 5 observed values.
    PII columns are redacted in the display values only.
    """

    seen: set[str] = set()
    rows: list[dict[str, Any]] = []
    policy_patterns = _workspace_redaction_patterns(workspace_path)
    for kpi in mapping.get("kpis", []):
        for feature in kpi.get("features", []):
            name = str(feature.get("feature") or "")
            if not name or name in seen:
                continue
            for source in feature.get("source_columns") or []:
                column = str(source.get("column") or "")
                if not column:
                    continue
                samples = list(
                    source.get("sample_values")
                    or (source.get("value_profile") or {}).get("sample_values")
                    or []
                )[:5]
                if not samples:
                    continue
                seen.add(name)
                display_samples = redact_sample_values(
                    column, samples, patterns=policy_patterns
                )
                rows.append(
                    {
                        "feature": name,
                        "column": f"{dataset_display_stem(str(source.get('dataset') or ''))}.{column}",
                        "first_samples": [str(value) for value in display_samples],
                        "redacted": is_pii_column(column, patterns=policy_patterns),
                    }
                )
                break  # one row per feature
    return rows


def _executable_sql_for_option(option: dict[str, Any], repo_root: Path) -> tuple[str, list[Path]]:
    """Return (sql, dataset_paths) for previewing an option.

    Physical-column options become a small SELECT against ``read_csv_auto``.
    Derived options compose the formula's input columns. Both cap with LIMIT 5.
    Returns ``("", [])`` when no executable SQL can be derived.
    """

    physical = option.get("physical_column_option") or {}
    if physical:
        dataset = str(physical.get("dataset") or "")
        column = str(physical.get("column") or "")
        if not dataset or not column:
            return "", []
        path = Path(dataset)
        if not path.is_absolute():
            path = repo_root / path
        sql = (
            f'SELECT "{column}"\n'
            f"FROM read_csv_auto('{dataset}')\n"
            f"WHERE \"{column}\" IS NOT NULL\n"
            f"LIMIT 5;"
        )
        return sql, [path]

    derived = option.get("derived_feature_option") or {}
    if derived:
        formula = str(derived.get("formula") or "")
        inputs = [col for col in derived.get("input_columns") or [] if isinstance(col, dict)]
        if not formula or not inputs:
            return "", []
        first = inputs[0]
        dataset = str(first.get("dataset") or "")
        if not dataset:
            return "", []
        path = Path(dataset)
        if not path.is_absolute():
            path = repo_root / path
        # Map "patients.DOB" -> '"DOB"' inside the formula (DuckDB sees the
        # CSV as a single unqualified table when read via read_csv_auto).
        sql_formula = formula
        for col in inputs:
            qualified = f"{dataset_display_stem(str(col.get('dataset') or ''))}.{col.get('column') or ''}"
            bare = str(col.get("column") or "")
            if qualified and bare:
                sql_formula = sql_formula.replace(qualified, f'"{bare}"')
        select_cols = ", ".join(f'"{str(c.get("column") or "")}"' for c in inputs if c.get("column"))
        derived_name = str(derived.get("derived_column_name") or "derived")
        sql = (
            f"SELECT {select_cols}, {sql_formula} AS {derived_name}\n"
            f"FROM read_csv_auto('{dataset}')\n"
            f"LIMIT 5;"
        )
        return sql, [path]

    return "", []


def _execute_option_preview(
    option: dict[str, Any],
    workspace_path: Path,
    repo_root: Path,
) -> dict[str, Any] | None:
    """Run an option's preview SQL (cache-first) and return a PreviewResult dict.

    Returns None when no executable SQL can be derived for the option (e.g.
    the ``custom`` placeholder option, or a CLI-agent-proposal option whose
    SQL isn't decided yet).

    Result rows have PII columns redacted for display.
    """

    sql, dataset_paths = _executable_sql_for_option(option, repo_root)
    if not sql:
        return None
    cache_key = compute_preview_cache_key(sql, dataset_paths)
    cached = load_cached_preview(workspace_path, cache_key)
    if cached is not None:
        return cached
    result = execute_preview(
        sql=sql,
        repo_root=repo_root,
        dataset_paths=dataset_paths,
    )
    payload = result.summary()
    if payload.get("status") == "ok" and payload.get("rows"):
        # PII redaction first, then injection neutralization: sample values
        # are untrusted workspace data rendered into an LLM-facing panel.
        payload["rows"] = neutralize_rows(
            redact_rows(
                payload["rows"],
                patterns=_workspace_redaction_patterns(workspace_path),
            )
        )
    if payload.get("status") == "ok":
        save_cached_preview(workspace_path, cache_key, payload)
    return payload


def _attach_preview_sections(
    panel: dict[str, Any],
    mapping: dict[str, Any],
    workspace_path: Path,
    repo_root: Path,
) -> None:
    """Mutate ``panel`` to add feature_resolution_table, sample_evidence, and
    per-option executed_sample blocks. KPI-level preview is filled when the
    recommended option has an executable preview that succeeded."""

    panel["feature_resolution_table"] = _build_feature_resolution_table(mapping)
    panel["sample_evidence"] = _build_sample_evidence(mapping, workspace_path)
    recommended_option_id = str(panel.get("recommended_option_id") or "")
    recommended_preview: dict[str, Any] | None = None
    for option in panel.get("options") or []:
        preview = _execute_option_preview(option, workspace_path, repo_root)
        if preview is not None:
            option["executed_sample"] = preview
            if option.get("option_id") == recommended_option_id and preview.get("status") == "ok":
                recommended_preview = preview
    if recommended_preview is not None:
        kpi_ids = panel.get("applies_to_kpis") or []
        kpi_id = kpi_ids[0] if kpi_ids else ""
        panel["kpi_preview"] = {
            "kpi_id": kpi_id,
            "assumed_option_id": recommended_option_id,
            "caption": (
                "Preview shows the recommended option's derivation/mapping over the "
                "first 5 rows. The full KPI SQL is generated only after every blocker "
                "is resolved."
            ),
            "preview_result": recommended_preview,
        }
